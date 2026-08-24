"""Tools the AI assistant can call.

Coverage now matches n8n's Google nodes one-for-one:
  Gmail   — read/search, get-full, send, reply, draft, modify-labels, list-labels, trash
  Calendar — list, get, create, update, delete, free/busy availability
  Tasks   — list lists, list, get, create, update (incl. mark-done), delete
  Contacts (People API) — list/search, get, create, update
  Memory  — vector recall over long-term notes

Plus the original Kin-local Supabase tasks/contacts so the agent can distinguish
between the user's *personal* lists in Kin and their Google data.
"""

from __future__ import annotations

import logging
import os
import re
import urllib.parse
from datetime import datetime, timedelta, timezone
from typing import Any, Optional

import httpx
from google.genai import types as genai_types

from plugins import doc_rag
from plugins import google_integrations as g
from plugins import memory as mem
from plugins import microsoft_integrations as ms
from plugins import notifications as notify

logger = logging.getLogger("kin.tools")


# ---------------------------------------------------------------------------
# Function declarations
# ---------------------------------------------------------------------------


def _schema(t: genai_types.Type, description: str = "") -> genai_types.Schema:
    return genai_types.Schema(type=t, description=description)


def _string_array(description: str) -> genai_types.Schema:
    return genai_types.Schema(
        type=genai_types.Type.ARRAY,
        description=description,
        items=genai_types.Schema(type=genai_types.Type.STRING),
    )


DECLARATIONS: list[genai_types.FunctionDeclaration] = [
    # -------- Email follow-ups ---------------------------------------------
    genai_types.FunctionDeclaration(
        name="send_followup_nudge",
        description=(
            "Send a polite one-line follow-up nudge on a Kin-sent email thread "
            "that's gone quiet. ONLY call this after Kin itself proactively "
            "asked the user 'want me to send a follow-up nudge?' (from the "
            "scheduled email-followup check) AND the user just said yes — that "
            "prior message includes a short reference code like '(ref: a1b2c3d4)' "
            "which you MUST pass as nudge_ref. Never call this speculatively or "
            "for a thread the user hasn't confirmed."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "nudge_ref": _schema(
                    genai_types.Type.STRING,
                    "The short reference code from Kin's own nudge-offer message, e.g. 'a1b2c3d4'.",
                ),
                "note": _schema(
                    genai_types.Type.STRING,
                    "Optional extra line the user wants added to the nudge (e.g. new info). Leave empty for a plain 'just checking in' nudge.",
                ),
            },
            required=["nudge_ref"],
        ),
    ),
    # -------- Gmail (read) ------------------------------------------------
    genai_types.FunctionDeclaration(
        name="read_gmail",
        description=(
            "List/search Gmail messages (metadata only — from, subject, snippet, "
            "date). Use this for inbox questions, security alerts, etc. Pass a "
            "Gmail-style query like 'from:google security', 'subject:invoice', "
            "'newer_than:30d'. You CAN search for ANY timeframe, including "
            "'yesterday' (e.g. 'newer_than:2d' or 'after:YYYY/MM/DD'). "
            "IMPORTANT: for time-based queries, set limit to 50. "
            "You have NO time limits. NEVER say you are limited to 7 days. "
            "If the user has connected extra Google accounts (Pro/Executive), "
            "results include messages from all of them — each has an 'account' "
            "field naming which one; mention it when results span more than one."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "query": _schema(genai_types.Type.STRING, "Gmail search query. Empty = recent inbox (excludes promotions/social)."),
                "limit": _schema(genai_types.Type.INTEGER, "Max emails (default 10, max 50). Use 30+ for multi-day queries."),
            },
        ),
    ),
    genai_types.FunctionDeclaration(
        name="get_gmail_message",
        description=(
            "Fetch one email's FULL content (headers + body, up to ~8000 chars). "
            "Call after read_gmail when the user wants details, a summary, or to "
            "draft a reply."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "message_id": _schema(genai_types.Type.STRING, "Gmail message ID."),
            },
            required=["message_id"],
        ),
    ),
    # -------- Gmail (write) -----------------------------------------------
    genai_types.FunctionDeclaration(
        name="send_email",
        description=(
            "Send a new email from the user's Gmail. ALWAYS confirm the recipient "
            "and content with the user first if not perfectly clear. To attach a "
            "file the user previously uploaded/indexed (e.g. their resume for a "
            "job application), pass its filename as attachment_file_name — Kin "
            "will attach the actual file, not just paste its text into the body. "
            "A signature is appended automatically if the user has one configured; "
            "never write your own signature/sign-off in the body."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "to": _string_array("Recipient email addresses."),
                "subject": _schema(genai_types.Type.STRING, "Subject line."),
                "body": _schema(genai_types.Type.STRING, "Plain-text body."),
                "cc": _string_array("Optional CC addresses."),
                "bcc": _string_array("Optional BCC addresses."),
                "attachment_file_name": _schema(
                    genai_types.Type.STRING,
                    "Optional: filename (or partial match) of a previously indexed document to attach.",
                ),
            },
            required=["to", "subject", "body"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="reply_email",
        description=(
            "Reply to an existing email, preserving the thread. Confirm content "
            "with the user before sending. To attach a previously indexed file, "
            "pass attachment_file_name. A signature is appended automatically if "
            "configured; never write your own sign-off in the body."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "message_id": _schema(genai_types.Type.STRING, "The Gmail message ID being replied to."),
                "body": _schema(genai_types.Type.STRING, "Reply body, plain text."),
                "attachment_file_name": _schema(
                    genai_types.Type.STRING,
                    "Optional: filename (or partial match) of a previously indexed document to attach.",
                ),
            },
            required=["message_id", "body"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="draft_email",
        description=(
            "Save a draft instead of sending. Use when the user wants to review "
            "the email in Gmail before sending it themselves."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "to": _string_array("Recipient addresses."),
                "subject": _schema(genai_types.Type.STRING, "Subject line."),
                "body": _schema(genai_types.Type.STRING, "Draft body."),
                "cc": _string_array("Optional CC addresses."),
            },
            required=["to", "subject", "body"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="modify_email_labels",
        description=(
            "Add or remove Gmail labels on a message. Use this to mark read/"
            "unread (UNREAD label), archive (remove INBOX), or apply custom "
            "labels by ID. Call list_gmail_labels first if you don't know the IDs."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "message_id": _schema(genai_types.Type.STRING, "Gmail message ID."),
                "add": _string_array("Label IDs to add (e.g. ['UNREAD', 'STARRED'])."),
                "remove": _string_array("Label IDs to remove (e.g. ['UNREAD'] to mark read)."),
            },
            required=["message_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="trash_email",
        description="Move a Gmail message to Trash. Reversible from Gmail UI within 30 days.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "message_id": _schema(genai_types.Type.STRING, "Gmail message ID."),
            },
            required=["message_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="list_gmail_labels",
        description="List the user's Gmail labels (system + custom) with their IDs.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT, properties={}
        ),
    ),
    genai_types.FunctionDeclaration(
        name="create_gmail_label",
        description="Create a new custom Gmail label.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "name": _schema(genai_types.Type.STRING, "Label name (e.g. 'Receipts')."),
            },
            required=["name"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="delete_email_permanent",
        description=(
            "PERMANENTLY delete an email — bypasses Trash, irreversible. "
            "Only use when the user explicitly says 'permanently delete' or "
            "'destroy'. Default to trash_email otherwise."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "message_id": _schema(genai_types.Type.STRING, "Gmail message ID."),
            },
            required=["message_id"],
        ),
    ),
    # -------- Gmail drafts ------------------------------------------------
    genai_types.FunctionDeclaration(
        name="list_drafts",
        description="List the user's Gmail drafts (most recent first).",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "limit": _schema(genai_types.Type.INTEGER, "Max drafts (default 20)."),
            },
        ),
    ),
    genai_types.FunctionDeclaration(
        name="get_draft",
        description="Get details of a specific Gmail draft.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "draft_id": _schema(genai_types.Type.STRING, "Draft ID."),
            },
            required=["draft_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="delete_draft",
        description="Delete a Gmail draft.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "draft_id": _schema(genai_types.Type.STRING, "Draft ID."),
            },
            required=["draft_id"],
        ),
    ),
    # -------- Gmail threads -----------------------------------------------
    genai_types.FunctionDeclaration(
        name="list_email_threads",
        description=(
            "List Gmail conversation threads (each thread is a group of replies). "
            "Use when the user asks about 'conversations' or 'threads' specifically."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "query": _schema(genai_types.Type.STRING, "Gmail search query (same syntax as read_gmail)."),
                "limit": _schema(genai_types.Type.INTEGER, "Max threads (default 20)."),
            },
        ),
    ),
    genai_types.FunctionDeclaration(
        name="get_email_thread",
        description=(
            "Fetch a full Gmail thread with all messages in chronological order. "
            "Use for 'show me my conversation with X' style questions."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "thread_id": _schema(genai_types.Type.STRING, "Gmail thread ID."),
            },
            required=["thread_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="reply_to_thread",
        description=(
            "Reply to the last message in a Gmail thread. Use this when the user "
            "references 'the thread' or 'the conversation' rather than a single "
            "message. Confirm content before sending. To attach a previously "
            "indexed file, pass attachment_file_name. A signature is appended "
            "automatically if configured; never write your own sign-off in the body."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "thread_id": _schema(genai_types.Type.STRING, "Gmail thread ID."),
                "body": _schema(genai_types.Type.STRING, "Reply body."),
                "attachment_file_name": _schema(
                    genai_types.Type.STRING,
                    "Optional: filename (or partial match) of a previously indexed document to attach.",
                ),
            },
            required=["thread_id", "body"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="trash_email_thread",
        description="Move an entire Gmail conversation thread to Trash.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "thread_id": _schema(genai_types.Type.STRING, "Gmail thread ID."),
            },
            required=["thread_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="modify_thread_labels",
        description=(
            "Add or remove labels on an entire thread at once. Useful for "
            "archiving (remove INBOX) or labeling a whole conversation."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "thread_id": _schema(genai_types.Type.STRING, "Gmail thread ID."),
                "add": _string_array("Label IDs to add."),
                "remove": _string_array("Label IDs to remove."),
            },
            required=["thread_id"],
        ),
    ),
    # -------- Calendar ----------------------------------------------------
    genai_types.FunctionDeclaration(
        name="read_calendar",
        description=(
            "List upcoming Google Calendar events. Use for schedule/availability "
            "questions. If the user has connected extra Google accounts "
            "(Pro/Executive), events include all of them — each has an 'account' "
            "field naming which one; mention it when results span more than one."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "days_ahead": _schema(
                    genai_types.Type.INTEGER,
                    "How many days ahead to look (default 1 = today only, max 30).",
                ),
            },
        ),
    ),
    genai_types.FunctionDeclaration(
        name="get_calendar_event",
        description="Fetch details of one calendar event by ID.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "event_id": _schema(genai_types.Type.STRING, "Calendar event ID."),
            },
            required=["event_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="create_calendar_event",
        description=(
            "Create a new Google Calendar event. Times should be ISO 8601 in the "
            "user's local timezone (e.g. '2026-05-12T15:00:00')."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "summary": _schema(genai_types.Type.STRING, "Event title."),
                "start": _schema(genai_types.Type.STRING, "ISO 8601 start datetime."),
                "end": _schema(genai_types.Type.STRING, "ISO 8601 end datetime."),
                "description": _schema(genai_types.Type.STRING, "Optional event description."),
                "location": _schema(genai_types.Type.STRING, "Optional physical/virtual location."),
                "attendees": _string_array("Optional list of attendee email addresses."),
                "all_day": _schema(genai_types.Type.BOOLEAN, "True for all-day events; start/end then become dates."),
            },
            required=["summary", "start", "end"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="update_calendar_event",
        description=(
            "Modify an existing Google Calendar event. IMPORTANT: Before calling this, "
            "ALWAYS call read_calendar first to find the event by its title/summary and "
            "get the correct event_id. Do NOT reuse event IDs from previous conversation "
            "turns — they may be stale or incorrect. Only pass fields you want to change."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "event_id": _schema(genai_types.Type.STRING, "Event to update."),
                "summary": _schema(genai_types.Type.STRING, "New title."),
                "start": _schema(genai_types.Type.STRING, "New ISO 8601 start."),
                "end": _schema(genai_types.Type.STRING, "New ISO 8601 end."),
                "description": _schema(genai_types.Type.STRING, "New description."),
                "location": _schema(genai_types.Type.STRING, "New location."),
                "attendees": _string_array("New full attendee list (replaces existing)."),
            },
            required=["event_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="delete_calendar_event",
        description="Cancel/delete a calendar event by ID.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "event_id": _schema(genai_types.Type.STRING, "Event ID."),
            },
            required=["event_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="check_calendar_availability",
        description=(
            "Free/busy query — returns intervals when the user is busy in a "
            "given window. Useful for 'am I free Tuesday at 3pm?'."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "start": _schema(genai_types.Type.STRING, "ISO 8601 start of the window."),
                "end": _schema(genai_types.Type.STRING, "ISO 8601 end of the window."),
            },
            required=["start", "end"],
        ),
    ),
    # -------- Google Tasks ------------------------------------------------
    genai_types.FunctionDeclaration(
        name="list_google_task_lists",
        description="List the user's Google Tasks task lists with their IDs.",
        parameters=genai_types.Schema(type=genai_types.Type.OBJECT, properties={}),
    ),
    genai_types.FunctionDeclaration(
        name="get_google_task",
        description="Get full details of a single Google Tasks task by ID.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "task_id": _schema(genai_types.Type.STRING, "Task ID."),
                "task_list_id": _schema(genai_types.Type.STRING, "Task list ID (default @default)."),
            },
            required=["task_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="list_google_tasks",
        description=(
            "List tasks from a Google Tasks list. Defaults to the user's primary "
            "list. Use this for Google Tasks specifically; use read_tasks for "
            "Kin-local tasks stored in the dashboard."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "task_list_id": _schema(genai_types.Type.STRING, "Task list ID (default: @default)."),
                "show_completed": _schema(genai_types.Type.BOOLEAN, "Include completed tasks."),
                "limit": _schema(genai_types.Type.INTEGER, "Max tasks (default 50)."),
            },
        ),
    ),
    genai_types.FunctionDeclaration(
        name="create_google_task",
        description="Create a new task in Google Tasks.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "title": _schema(genai_types.Type.STRING, "Task title."),
                "notes": _schema(genai_types.Type.STRING, "Optional details."),
                "due": _schema(genai_types.Type.STRING, "Optional RFC 3339 due date."),
                "task_list_id": _schema(genai_types.Type.STRING, "Task list ID (default @default)."),
            },
            required=["title"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="update_google_task",
        description="Update a Google Tasks task — title, notes, due date, or mark completed.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "task_id": _schema(genai_types.Type.STRING, "Task ID."),
                "title": _schema(genai_types.Type.STRING, "New title."),
                "notes": _schema(genai_types.Type.STRING, "New notes."),
                "due": _schema(genai_types.Type.STRING, "New RFC 3339 due date."),
                "completed": _schema(genai_types.Type.BOOLEAN, "True to mark done, false to reopen."),
                "task_list_id": _schema(genai_types.Type.STRING, "Task list ID (default @default)."),
            },
            required=["task_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="delete_google_task",
        description="Delete a Google Tasks task.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "task_id": _schema(genai_types.Type.STRING, "Task ID."),
                "task_list_id": _schema(genai_types.Type.STRING, "Task list ID (default @default)."),
            },
            required=["task_id"],
        ),
    ),
    # -------- Google Contacts (People API) --------------------------------
    genai_types.FunctionDeclaration(
        name="list_google_contacts",
        description=(
            "List or search the user's Google Contacts. Pass a query to filter "
            "by name/email/etc; empty query lists most-recently-modified contacts."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "query": _schema(genai_types.Type.STRING, "Search query, or empty for recent contacts."),
                "limit": _schema(genai_types.Type.INTEGER, "Max contacts (default 30)."),
            },
        ),
    ),
    genai_types.FunctionDeclaration(
        name="get_google_contact",
        description="Fetch a Google contact by resourceName (e.g. 'people/c1234567').",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "resource_name": _schema(genai_types.Type.STRING, "People API resourceName."),
            },
            required=["resource_name"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="create_google_contact",
        description="Create a new Google contact.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "name": _schema(genai_types.Type.STRING, "Display name / given name."),
                "email": _schema(genai_types.Type.STRING, "Optional email."),
                "phone": _schema(genai_types.Type.STRING, "Optional phone."),
                "company": _schema(genai_types.Type.STRING, "Optional company."),
                "notes": _schema(genai_types.Type.STRING, "Optional notes."),
            },
            required=["name"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="update_google_contact",
        description="Update fields on an existing Google contact.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "resource_name": _schema(genai_types.Type.STRING, "People API resourceName."),
                "name": _schema(genai_types.Type.STRING, "New display name."),
                "email": _schema(genai_types.Type.STRING, "New email (replaces existing)."),
                "phone": _schema(genai_types.Type.STRING, "New phone (replaces existing)."),
                "company": _schema(genai_types.Type.STRING, "New company."),
                "notes": _schema(genai_types.Type.STRING, "New notes."),
            },
            required=["resource_name"],
        ),
    ),
    # -------- Kin-local Supabase tasks/contacts ---------------------------
    genai_types.FunctionDeclaration(
        name="read_tasks",
        description=(
            "Read Kin-local tasks (NOT Google Tasks) — these are the tasks the "
            "user manages on the Kin dashboard. Use this when the user says "
            "'my Kin tasks' or just 'tasks' without specifying."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "status": _schema(genai_types.Type.STRING, "todo, in_progress, done, or open (default open)."),
                "limit": _schema(genai_types.Type.INTEGER, "Max tasks (default 20)."),
            },
        ),
    ),
    genai_types.FunctionDeclaration(
        name="read_contacts",
        description=(
            "Search Kin-local contacts (NOT Google Contacts). Use for the "
            "address book the user manages on the Kin dashboard."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "query": _schema(genai_types.Type.STRING, "Name, email, or company substring."),
                "limit": _schema(genai_types.Type.INTEGER, "Max contacts (default 10)."),
            },
        ),
    ),
    # -------- Drive / Docs / Sheets / Slides ------------------------------
    genai_types.FunctionDeclaration(
        name="search_documents",
        description=(
            "Semantic search across the user's INDEXED documents — files from "
            "Google Drive, OneDrive, AND any PDFs/DOCX/TXT the user has uploaded "
            "directly through web chat. Use this for "
            "a SPECIFIC question a file might answer — a policy, a figure, a "
            "clause, a fact — not for 'summarize this' or 'what's in this "
            "document' style requests, which need the WHOLE document (use "
            "read_full_document for those instead — semantic search only "
            "returns the few chunks most similar to your exact query wording, "
            "which is unreliable for broad requests since a word like "
            "'summary' doesn't semantically resemble the document's actual "
            "content). Returns the most relevant text chunks with their source "
            "filenames so you can cite them in your reply. Always cite the "
            "filename(s) you used. If this returns 0 results, the file may not "
            "be indexed yet — try list_drive_files to find it, then suggest "
            "the user indexes it at /dashboard/documents (or asks them to "
            "re-upload on chat). NEVER say you cannot read a file type."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "query": _schema(genai_types.Type.STRING, "Natural-language question to search for."),
                "limit": _schema(genai_types.Type.INTEGER, "Max chunks to return (default 8, max 15)."),
            },
            required=["query"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="read_full_document",
        description=(
            "Fetch the COMPLETE indexed text of ONE document by filename (all "
            "its chunks, concatenated in order) — use this for 'summarize "
            "this', 'what does this say overall', 'give me an overview of X' "
            "style requests, on the FIRST attempt, instead of retrying "
            "search_documents with rephrased queries. Especially useful right "
            "after the user just uploaded/indexed a file and asks about it "
            "generally. Matches filename by partial/case-insensitive text; if "
            "several files match, it reads the most recently indexed one and "
            "says so."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "file_name": _schema(
                    genai_types.Type.STRING,
                    "Filename or partial filename, e.g. 'Tharindu_Resume.pdf' or just 'resume'.",
                ),
            },
            required=["file_name"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="list_drive_files",
        description=(
            "List/search files in the user's Google Drive (not the indexed RAG "
            "store — this is the LIVE Drive). Use this when the user wants to "
            "find a file by name or browse a folder."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "query": _schema(genai_types.Type.STRING, "Filename substring to filter by."),
                "folder_id": _schema(genai_types.Type.STRING, "Optional Drive folder ID to scope the list."),
                "limit": _schema(genai_types.Type.INTEGER, "Max files (default 30)."),
            },
        ),
    ),
    genai_types.FunctionDeclaration(
        name="index_drive_file",
        description=(
            "Index a SINGLE Google Drive file for RAG so the user can ask "
            "questions about its contents. Workflow: call list_drive_files "
            "FIRST to find the file by name and get its `file_id`, THEN call "
            "this with that id. After indexing finishes, you can call "
            "search_documents to answer questions grounded in the file. "
            "Tell the user 'Indexing <filename>…' before calling, and report "
            "the chunk_count after. Supports Docs, Sheets, Slides, PDF, DOCX, "
            "TXT, MD, CSV."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "file_id": _schema(genai_types.Type.STRING, "Drive file ID (from list_drive_files)."),
            },
            required=["file_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="index_drive_folder",
        description=(
            "Index EVERY supported file in a Google Drive folder for RAG. "
            "Workflow: use list_drive_files with the folder name to find its "
            "folder_id (a folder has mime_type 'application/vnd.google-apps.folder'), "
            "THEN call this. Non-recursive (only direct children). After "
            "indexing finishes you can call search_documents to answer "
            "questions grounded in the folder contents."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "folder_id": _schema(genai_types.Type.STRING, "Drive folder ID."),
                "max_files": _schema(genai_types.Type.INTEGER, "Cap files to index (default 25, max 50)."),
            },
            required=["folder_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="index_onedrive_file",
        description=(
            "Index a SINGLE OneDrive file for RAG so the user can ask questions "
            "about its contents. Workflow: call list_onedrive_files first to "
            "find the file by name and get its item_id, THEN call this with "
            "that id. Supports PDF, DOCX, TXT, MD, CSV."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "file_id": _schema(genai_types.Type.STRING, "OneDrive item ID (from list_onedrive_files)."),
            },
            required=["file_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="index_onedrive_folder",
        description=(
            "Index every supported file in a OneDrive folder (non-recursive). "
            "Workflow: list_onedrive_files with the folder name first to get "
            "its item_id, THEN call this."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "folder_id": _schema(genai_types.Type.STRING, "OneDrive folder item ID."),
                "max_files": _schema(genai_types.Type.INTEGER, "Cap files to index (default 25, max 50)."),
            },
            required=["folder_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="get_pdf_info",
        description=(
            "Get metadata about a PDF (page count, size, whether it's scanned). "
            "Works on both Google Drive and OneDrive PDFs. Workflow: find the "
            "file with list_drive_files / list_onedrive_files first, then call "
            "this with its id. Returns {page_count, size_bytes, "
            "has_extractable_text}. If has_extractable_text=false, the PDF is "
            "scanned (image-only) and indexing won't work."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "file_id": _schema(genai_types.Type.STRING, "Drive/OneDrive file id."),
                "source": _schema(genai_types.Type.STRING, "'gdrive' (default) or 'onedrive'."),
            },
            required=["file_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="get_spreadsheet_info",
        description=(
            "Get sheet names, sheet count, and row/column counts for an "
            "uploaded .xlsx/.xls spreadsheet. DO NOT use search_documents for "
            "this — semantic chunk search cannot answer structural questions "
            "like 'how many sheets/tabs does this file have' or 'what are "
            "the sheet names', so call get_spreadsheet_info instead, as the "
            "FIRST and ONLY tool call, the moment you see this kind of "
            "question. If the user names the file (e.g. 'the indexed "
            "200097E.xlsx'), pass that as `file_name` directly — do NOT call "
            "list_drive_files/list_onedrive_files first, this tool looks the "
            "file up itself from what's already indexed. Only pass `file_id`"
            "+`source` if you already have them from an earlier tool result "
            "in this conversation. Not for native Google Sheets — use "
            "list_sheet_tabs for those."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "file_name": _schema(genai_types.Type.STRING, "Filename (or substring) of an already-indexed spreadsheet, e.g. '200097E.xlsx'."),
                "file_id": _schema(genai_types.Type.STRING, "Drive/OneDrive file id, if already known."),
                "source": _schema(genai_types.Type.STRING, "'gdrive' or 'onedrive' — only used together with file_id."),
            },
        ),
    ),
    genai_types.FunctionDeclaration(
        name="send_file_to_user",
        description=(
            "Send a file from Google Drive or OneDrive directly to the user in "
            "this chat. "
            "Returns a "
            "download URL the user can click. Use this when the user says "
            "'send me the PDF', 'download X', 'share with me', 'give me the "
            "file' — they mean themselves, NOT another recipient. Find the "
            "file with list_drive_files / list_onedrive_files first to get "
            "the id, then call this. Max file size 50 MB."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "file_id": _schema(genai_types.Type.STRING, "Drive/OneDrive file id."),
                "source": _schema(genai_types.Type.STRING, "'gdrive' (default) or 'onedrive'."),
            },
            required=["file_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="read_google_doc",
        description=(
            "Read the FULL text of a Google Doc by its document ID. Use when the "
            "user references a specific doc and search_documents wasn't enough."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "document_id": _schema(genai_types.Type.STRING, "Google Doc document ID."),
            },
            required=["document_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="read_google_sheet",
        description=(
            "Read a range of cells from a Google Sheet. Default range covers "
            "A1:Z100; pass a specific A1 range like 'Sheet2!A1:E50' to scope it."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "spreadsheet_id": _schema(genai_types.Type.STRING, "Spreadsheet ID."),
                "range_a1": _schema(genai_types.Type.STRING, "A1 range, e.g. 'Sheet1!A1:Z100'."),
            },
            required=["spreadsheet_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="list_sheet_tabs",
        description="List the tabs (sheets) in a Google Sheets spreadsheet.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "spreadsheet_id": _schema(genai_types.Type.STRING, "Spreadsheet ID."),
            },
            required=["spreadsheet_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="read_google_slides",
        description="Read the full text content of a Google Slides presentation.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "presentation_id": _schema(genai_types.Type.STRING, "Presentation ID."),
            },
            required=["presentation_id"],
        ),
    ),
    # -------- Drive writes ------------------------------------------------
    genai_types.FunctionDeclaration(
        name="save_to_drive",
        description=(
            "Save text content as a new file in the user's Drive (e.g. notes, "
            "meeting summaries). Confirm name + content with the user first."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "name": _schema(genai_types.Type.STRING, "Filename including extension (e.g. 'Notes.md')."),
                "text": _schema(genai_types.Type.STRING, "File contents."),
                "mime_type": _schema(genai_types.Type.STRING, "MIME type (default text/plain)."),
                "parent_folder_id": _schema(genai_types.Type.STRING, "Optional parent folder ID."),
            },
            required=["name", "text"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="create_drive_folder",
        description="Create a new folder in Drive.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "name": _schema(genai_types.Type.STRING, "Folder name."),
                "parent_folder_id": _schema(genai_types.Type.STRING, "Optional parent folder ID."),
            },
            required=["name"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="copy_drive_file",
        description="Make a copy of a Drive file.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "file_id": _schema(genai_types.Type.STRING, "Source file ID."),
                "name": _schema(genai_types.Type.STRING, "New name for the copy."),
                "parent_folder_id": _schema(genai_types.Type.STRING, "Optional destination folder."),
            },
            required=["file_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="move_drive_file",
        description="Move a Drive file to a different folder.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "file_id": _schema(genai_types.Type.STRING, "File to move."),
                "new_parent_folder_id": _schema(genai_types.Type.STRING, "Destination folder ID."),
            },
            required=["file_id", "new_parent_folder_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="rename_drive_file",
        description="Rename a Drive file or folder, or update its description.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "file_id": _schema(genai_types.Type.STRING, "File or folder ID."),
                "name": _schema(genai_types.Type.STRING, "New name."),
                "description": _schema(genai_types.Type.STRING, "New description."),
            },
            required=["file_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="delete_drive_item",
        description=(
            "Move a Drive file or folder to Trash (or permanently delete). "
            "Default is Trash — only set permanent=true on explicit user request."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "file_id": _schema(genai_types.Type.STRING, "File or folder ID."),
                "permanent": _schema(genai_types.Type.BOOLEAN, "True = bypass Trash (irreversible)."),
            },
            required=["file_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="share_drive_item",
        description=(
            "Share a Drive file or folder with someone by email. role: reader, "
            "commenter, or writer. Confirm with the user before sharing externally."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "file_id": _schema(genai_types.Type.STRING, "File or folder to share."),
                "email": _schema(genai_types.Type.STRING, "Person's email address."),
                "role": _schema(genai_types.Type.STRING, "reader | commenter | writer."),
            },
            required=["file_id", "email"],
        ),
    ),
    # -------- Docs writes -------------------------------------------------
    genai_types.FunctionDeclaration(
        name="create_google_doc",
        description="Create a new empty Google Doc with the given title.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "title": _schema(genai_types.Type.STRING, "Doc title."),
            },
            required=["title"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="append_to_google_doc",
        description="Append text to the end of an existing Google Doc.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "document_id": _schema(genai_types.Type.STRING, "Doc ID."),
                "text": _schema(genai_types.Type.STRING, "Text to append."),
            },
            required=["document_id", "text"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="replace_text_in_doc",
        description="Find-and-replace text inside a Google Doc.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "document_id": _schema(genai_types.Type.STRING, "Doc ID."),
                "find": _schema(genai_types.Type.STRING, "Text to find."),
                "replace": _schema(genai_types.Type.STRING, "Replacement text."),
            },
            required=["document_id", "find", "replace"],
        ),
    ),
    # -------- Slides writes ----------------------------------------------
    genai_types.FunctionDeclaration(
        name="replace_text_in_slides",
        description=(
            "Find-and-replace text across all slides in a presentation. Great "
            "for template substitution (e.g. replace {{name}} with the user's name)."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "presentation_id": _schema(genai_types.Type.STRING, "Presentation ID."),
                "find": _schema(genai_types.Type.STRING, "Text/placeholder to find."),
                "replace": _schema(genai_types.Type.STRING, "Replacement text."),
            },
            required=["presentation_id", "find", "replace"],
        ),
    ),
    # -------- Sheets writes ----------------------------------------------
    genai_types.FunctionDeclaration(
        name="create_spreadsheet",
        description="Create a new Google Sheets spreadsheet.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "title": _schema(genai_types.Type.STRING, "Spreadsheet title."),
                "sheet_titles": _string_array("Optional list of initial tab names."),
            },
            required=["title"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="add_sheet_tab",
        description="Add a new tab (sheet) to an existing spreadsheet.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "spreadsheet_id": _schema(genai_types.Type.STRING, "Spreadsheet ID."),
                "title": _schema(genai_types.Type.STRING, "New tab name."),
            },
            required=["spreadsheet_id", "title"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="append_sheet_row",
        description=(
            "Append one or more rows to the end of a Google Sheet range. "
            "Each row is a list of cell values. Useful for 'log this entry to my "
            "sheet' style requests."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "spreadsheet_id": _schema(genai_types.Type.STRING, "Spreadsheet ID."),
                "range_a1": _schema(genai_types.Type.STRING, "A1 range, e.g. 'Sheet1!A:E'."),
                "rows": genai_types.Schema(
                    type=genai_types.Type.ARRAY,
                    description="Rows to append. Each row is an array of cell values.",
                    items=genai_types.Schema(
                        type=genai_types.Type.ARRAY,
                        items=genai_types.Schema(type=genai_types.Type.STRING),
                    ),
                ),
            },
            required=["spreadsheet_id", "range_a1", "rows"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="update_sheet_values",
        description="Write values to a specific range in a Google Sheet (overwrites existing cells).",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "spreadsheet_id": _schema(genai_types.Type.STRING, "Spreadsheet ID."),
                "range_a1": _schema(genai_types.Type.STRING, "A1 range, e.g. 'Sheet1!A1:C3'."),
                "rows": genai_types.Schema(
                    type=genai_types.Type.ARRAY,
                    description="2D array of values to write.",
                    items=genai_types.Schema(
                        type=genai_types.Type.ARRAY,
                        items=genai_types.Schema(type=genai_types.Type.STRING),
                    ),
                ),
            },
            required=["spreadsheet_id", "range_a1", "rows"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="clear_sheet_range",
        description="Clear all values in a Google Sheet range (cells become empty).",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "spreadsheet_id": _schema(genai_types.Type.STRING, "Spreadsheet ID."),
                "range_a1": _schema(genai_types.Type.STRING, "A1 range to clear."),
            },
            required=["spreadsheet_id", "range_a1"],
        ),
    ),
    # -------- Microsoft Outlook ------------------------------------------
    genai_types.FunctionDeclaration(
        name="read_outlook",
        description=(
            "List recent emails from the user's Microsoft Outlook inbox. "
            "Use this when the user asks about their Outlook mail. "
            "Optional Graph search query (e.g. 'from:boss' or 'invoice'). "
            "IMPORTANT: For time-based requests (like 'last 30 days' or 'yesterday'), "
            "DO NOT use a date query. Leave the query EMPTY, set limit to 50, "
            "and ALWAYS pass the `max_days_old` parameter. "
            "This ensures emails are filtered accurately by the backend."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "query": _schema(genai_types.Type.STRING, "Search query (leave empty for date ranges)."),
                "limit": _schema(genai_types.Type.INTEGER, "Max emails (default 15, max 50)."),
                "folder": _schema(genai_types.Type.STRING, "Folder name (default 'Inbox')."),
                "max_days_old": _schema(genai_types.Type.INTEGER, "Filter to emails received in the last N days (e.g. 30, 7, 1)."),
            },
        ),
    ),
    genai_types.FunctionDeclaration(
        name="get_outlook_message",
        description="Fetch one Outlook email's full body + metadata by ID.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "message_id": _schema(genai_types.Type.STRING, "Outlook message ID."),
            },
            required=["message_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="send_outlook_email",
        description=(
            "Send an email through Microsoft Outlook. Confirm recipient and "
            "content with the user first."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "to": _string_array("Recipient email addresses."),
                "subject": _schema(genai_types.Type.STRING, "Subject line."),
                "body": _schema(genai_types.Type.STRING, "Plain-text body."),
                "cc": _string_array("Optional CC addresses."),
                "bcc": _string_array("Optional BCC addresses."),
            },
            required=["to", "subject", "body"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="reply_outlook_email",
        description="Reply to an existing Outlook email by ID, preserving the thread.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "message_id": _schema(genai_types.Type.STRING, "Outlook message ID."),
                "body": _schema(genai_types.Type.STRING, "Reply body."),
            },
            required=["message_id", "body"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="mark_outlook_read",
        description="Mark an Outlook message as read or unread.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "message_id": _schema(genai_types.Type.STRING, "Outlook message ID."),
                "is_read": _schema(genai_types.Type.BOOLEAN, "True=read, false=unread."),
            },
            required=["message_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="delete_outlook_message",
        description="Soft-delete an Outlook message (moves to Deleted Items).",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "message_id": _schema(genai_types.Type.STRING, "Outlook message ID."),
            },
            required=["message_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="move_outlook_message",
        description=(
            "Move an Outlook message to another folder. destination_folder can "
            "be a well-known name like 'archive' or a folder ID."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "message_id": _schema(genai_types.Type.STRING, "Outlook message ID."),
                "destination_folder": _schema(
                    genai_types.Type.STRING,
                    "Folder name (archive, junkemail, deleteditems, inbox) or folder ID.",
                ),
            },
            required=["message_id", "destination_folder"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="list_outlook_folders",
        description="List the user's Outlook mail folders with counts.",
        parameters=genai_types.Schema(type=genai_types.Type.OBJECT, properties={}),
    ),
    # -------- Microsoft OneDrive -----------------------------------------
    genai_types.FunctionDeclaration(
        name="list_onedrive_files",
        description=(
            "List or search files in OneDrive (live, not the RAG store). "
            "Used to find a OneDrive file by name."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "query": _schema(genai_types.Type.STRING, "Filename substring to search for."),
                "folder_id": _schema(genai_types.Type.STRING, "Optional OneDrive folder ID."),
                "limit": _schema(genai_types.Type.INTEGER, "Max files (default 30)."),
            },
        ),
    ),
    genai_types.FunctionDeclaration(
        name="get_onedrive_metadata",
        description="Get metadata for one OneDrive file by ID.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "item_id": _schema(genai_types.Type.STRING, "OneDrive item ID."),
            },
            required=["item_id"],
        ),
    ),
    # -------- Microsoft ToDo ---------------------------------------------
    genai_types.FunctionDeclaration(
        name="list_todo_lists",
        description="List the user's Microsoft ToDo task lists.",
        parameters=genai_types.Schema(type=genai_types.Type.OBJECT, properties={}),
    ),
    genai_types.FunctionDeclaration(
        name="list_todo_tasks",
        description=(
            "List tasks in Microsoft ToDo. Defaults to the user's primary list "
            "when list_id is omitted."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "list_id": _schema(genai_types.Type.STRING, "ToDo list ID (optional)."),
                "show_completed": _schema(genai_types.Type.BOOLEAN, "Include completed."),
                "limit": _schema(genai_types.Type.INTEGER, "Max tasks (default 50)."),
            },
        ),
    ),
    genai_types.FunctionDeclaration(
        name="create_todo_task",
        description="Create a new task in Microsoft ToDo.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "title": _schema(genai_types.Type.STRING, "Task title."),
                "body": _schema(genai_types.Type.STRING, "Optional details."),
                "due": _schema(genai_types.Type.STRING, "Optional ISO due datetime."),
                "importance": _schema(genai_types.Type.STRING, "low | normal | high."),
                "list_id": _schema(genai_types.Type.STRING, "Optional ToDo list ID."),
            },
            required=["title"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="update_todo_task",
        description="Update a Microsoft ToDo task — title, body, due, importance, or mark completed.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "task_id": _schema(genai_types.Type.STRING, "Task ID."),
                "list_id": _schema(genai_types.Type.STRING, "List ID containing the task."),
                "title": _schema(genai_types.Type.STRING, "New title."),
                "body": _schema(genai_types.Type.STRING, "New body."),
                "completed": _schema(genai_types.Type.BOOLEAN, "Mark completed/uncompleted."),
                "due": _schema(genai_types.Type.STRING, "ISO due datetime."),
                "importance": _schema(genai_types.Type.STRING, "low | normal | high."),
            },
            required=["task_id", "list_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="delete_todo_task",
        description="Delete a Microsoft ToDo task.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "task_id": _schema(genai_types.Type.STRING, "Task ID."),
                "list_id": _schema(genai_types.Type.STRING, "List ID containing the task."),
            },
            required=["task_id", "list_id"],
        ),
    ),
    # -------- OneDrive writes --------------------------------------------
    genai_types.FunctionDeclaration(
        name="save_to_onedrive",
        description=(
            "Save text content as a new file in the user's OneDrive. "
            "Confirm name + content with the user before saving."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "name": _schema(genai_types.Type.STRING, "Filename including extension."),
                "text": _schema(genai_types.Type.STRING, "File contents."),
                "mime_type": _schema(genai_types.Type.STRING, "MIME type (default text/plain)."),
                "parent_folder_id": _schema(genai_types.Type.STRING, "Optional parent folder ID."),
            },
            required=["name", "text"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="create_onedrive_folder",
        description="Create a new folder in OneDrive.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "name": _schema(genai_types.Type.STRING, "Folder name."),
                "parent_folder_id": _schema(genai_types.Type.STRING, "Optional parent."),
            },
            required=["name"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="rename_onedrive_item",
        description="Rename a OneDrive file or folder.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "item_id": _schema(genai_types.Type.STRING, "Item ID."),
                "new_name": _schema(genai_types.Type.STRING, "New name."),
            },
            required=["item_id", "new_name"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="move_onedrive_item",
        description="Move a OneDrive file or folder to another folder.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "item_id": _schema(genai_types.Type.STRING, "Item to move."),
                "new_parent_folder_id": _schema(genai_types.Type.STRING, "Destination folder ID."),
            },
            required=["item_id", "new_parent_folder_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="copy_onedrive_item",
        description="Copy a OneDrive file or folder.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "item_id": _schema(genai_types.Type.STRING, "Source item ID."),
                "new_name": _schema(genai_types.Type.STRING, "New name."),
                "parent_folder_id": _schema(genai_types.Type.STRING, "Optional destination folder."),
            },
            required=["item_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="delete_onedrive_item",
        description="Delete a OneDrive file or folder (moves to Recycle Bin).",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "item_id": _schema(genai_types.Type.STRING, "Item ID."),
            },
            required=["item_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="share_onedrive_item",
        description=(
            "Share a OneDrive item — either invite a specific email or "
            "(if email omitted) generate a sharing link. role: 'read' or 'write'."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "item_id": _schema(genai_types.Type.STRING, "Item ID."),
                "email": _schema(genai_types.Type.STRING, "Recipient email (optional)."),
                "role": _schema(genai_types.Type.STRING, "'read' or 'write'."),
            },
            required=["item_id"],
        ),
    ),
    # -------- Outlook drafts ---------------------------------------------
    genai_types.FunctionDeclaration(
        name="create_outlook_draft",
        description="Create a draft Outlook email — saved in Drafts, not sent.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "to": _string_array("Recipient email addresses."),
                "subject": _schema(genai_types.Type.STRING, "Subject."),
                "body": _schema(genai_types.Type.STRING, "Body."),
                "cc": _string_array("Optional CC."),
            },
            required=["to", "subject", "body"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="update_outlook_draft",
        description="Update an existing Outlook draft.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "draft_id": _schema(genai_types.Type.STRING, "Draft ID."),
                "subject": _schema(genai_types.Type.STRING, "New subject."),
                "body": _schema(genai_types.Type.STRING, "New body."),
                "to": _string_array("New recipient list."),
            },
            required=["draft_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="send_outlook_draft",
        description="Send a previously saved Outlook draft by ID.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "draft_id": _schema(genai_types.Type.STRING, "Draft ID."),
            },
            required=["draft_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="update_outlook_message",
        description=(
            "Update flags/categories on an Outlook message: mark read/unread, "
            "flag for follow-up, or set categories."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "message_id": _schema(genai_types.Type.STRING, "Message ID."),
                "is_read": _schema(genai_types.Type.BOOLEAN, "Mark read/unread."),
                "flag": _schema(
                    genai_types.Type.STRING,
                    "Flag status: notFlagged | flagged | complete.",
                ),
                "categories": _string_array("Categories to apply (replaces existing)."),
            },
            required=["message_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="list_outlook_folder_messages",
        description="List messages inside a specific Outlook folder by folder ID.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "folder_id": _schema(genai_types.Type.STRING, "Outlook folder ID."),
                "limit": _schema(genai_types.Type.INTEGER, "Max messages (default 25)."),
            },
            required=["folder_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="create_outlook_folder",
        description="Create a new Outlook mail folder.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "name": _schema(genai_types.Type.STRING, "Folder name."),
                "parent_folder_id": _schema(genai_types.Type.STRING, "Optional parent."),
            },
            required=["name"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="get_outlook_folder",
        description="Get details of an Outlook mail folder.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "folder_id": _schema(genai_types.Type.STRING, "Folder ID."),
            },
            required=["folder_id"],
        ),
    ),
    # -------- Outlook calendars + events ---------------------------------
    genai_types.FunctionDeclaration(
        name="list_outlook_calendars",
        description="List the user's Outlook calendars.",
        parameters=genai_types.Schema(type=genai_types.Type.OBJECT, properties={}),
    ),
    genai_types.FunctionDeclaration(
        name="create_outlook_calendar",
        description="Create a new Outlook calendar.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "name": _schema(genai_types.Type.STRING, "Calendar name."),
            },
            required=["name"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="get_outlook_calendar",
        description="Get details of one Outlook calendar by ID.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "calendar_id": _schema(genai_types.Type.STRING, "Calendar ID."),
            },
            required=["calendar_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="list_outlook_events",
        description=(
            "List Outlook calendar events. Optional days_ahead window similar to "
            "the Google Calendar tool."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "calendar_id": _schema(genai_types.Type.STRING, "Optional calendar ID (default: primary)."),
                "days_ahead": _schema(genai_types.Type.INTEGER, "Lookahead window (default 7)."),
                "limit": _schema(genai_types.Type.INTEGER, "Max events (default 25)."),
            },
        ),
    ),
    genai_types.FunctionDeclaration(
        name="get_outlook_event",
        description="Fetch one Outlook event.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "event_id": _schema(genai_types.Type.STRING, "Event ID."),
            },
            required=["event_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="create_outlook_event",
        description="Create an Outlook calendar event. Times in ISO 8601 in user's local timezone.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "subject": _schema(genai_types.Type.STRING, "Event title."),
                "start": _schema(genai_types.Type.STRING, "ISO 8601 start datetime."),
                "end": _schema(genai_types.Type.STRING, "ISO 8601 end datetime."),
                "body": _schema(genai_types.Type.STRING, "Optional description."),
                "location": _schema(genai_types.Type.STRING, "Optional location."),
                "attendees": _string_array("Optional attendee emails."),
                "is_all_day": _schema(genai_types.Type.BOOLEAN, "All-day event."),
                "calendar_id": _schema(genai_types.Type.STRING, "Optional non-default calendar."),
                "online_meeting": _schema(genai_types.Type.BOOLEAN, "Set true to create a Microsoft Teams online meeting and generate a join link."),
            },
            required=["subject", "start", "end"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="update_outlook_event",
        description="Modify an existing Outlook event. Only pass fields you want to change.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "event_id": _schema(genai_types.Type.STRING, "Event ID."),
                "subject": _schema(genai_types.Type.STRING, "New title."),
                "start": _schema(genai_types.Type.STRING, "New ISO 8601 start."),
                "end": _schema(genai_types.Type.STRING, "New ISO 8601 end."),
                "body": _schema(genai_types.Type.STRING, "New description."),
                "location": _schema(genai_types.Type.STRING, "New location."),
                "attendees": _string_array("New full attendee list."),
            },
            required=["event_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="delete_outlook_event",
        description="Cancel/delete an Outlook event by ID.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "event_id": _schema(genai_types.Type.STRING, "Event ID."),
            },
            required=["event_id"],
        ),
    ),
    # -------- Outlook contacts -------------------------------------------
    genai_types.FunctionDeclaration(
        name="list_outlook_contacts",
        description="List the user's Outlook contacts.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "limit": _schema(genai_types.Type.INTEGER, "Max contacts (default 50)."),
            },
        ),
    ),
    genai_types.FunctionDeclaration(
        name="get_outlook_contact",
        description="Fetch one Outlook contact by ID.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "contact_id": _schema(genai_types.Type.STRING, "Contact ID."),
            },
            required=["contact_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="create_outlook_contact",
        description="Create a new Outlook contact.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "name": _schema(genai_types.Type.STRING, "Display/given name."),
                "email": _schema(genai_types.Type.STRING, "Optional email."),
                "phone": _schema(genai_types.Type.STRING, "Optional phone."),
                "company": _schema(genai_types.Type.STRING, "Optional company."),
                "job_title": _schema(genai_types.Type.STRING, "Optional job title."),
            },
            required=["name"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="delete_outlook_contact",
        description="Delete an Outlook contact by ID.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "contact_id": _schema(genai_types.Type.STRING, "Contact ID."),
            },
            required=["contact_id"],
        ),
    ),
    # -------- Outlook attachments ----------------------------------------
    genai_types.FunctionDeclaration(
        name="list_outlook_attachments",
        description="List attachments on an Outlook message.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "message_id": _schema(genai_types.Type.STRING, "Message ID."),
            },
            required=["message_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="get_outlook_attachment",
        description="Get metadata for one Outlook attachment.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "message_id": _schema(genai_types.Type.STRING, "Message ID."),
                "attachment_id": _schema(genai_types.Type.STRING, "Attachment ID."),
            },
            required=["message_id", "attachment_id"],
        ),
    ),
    # -------- ToDo list create -------------------------------------------
    genai_types.FunctionDeclaration(
        name="create_todo_list",
        description="Create a new Microsoft ToDo task list.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "display_name": _schema(genai_types.Type.STRING, "List name."),
            },
            required=["display_name"],
        ),
    ),
    # -------- Meeting prep --------------------------------------------------
    genai_types.FunctionDeclaration(
        name="prepare_meeting_brief",
        description=(
            "Build a pre-meeting brief for one calendar event: pulls the "
            "event details, then for each attendee looks up their contact "
            "info and recent email history with the user. Call read_calendar "
            "or list_outlook_events FIRST to find the event_id if you don't "
            "already have it (e.g. for 'brief me for my next meeting', list "
            "today's events first, pick the next upcoming one, then call "
            "this with its id)."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "event_id": _schema(genai_types.Type.STRING, "The calendar event's id."),
                "source": genai_types.Schema(
                    type=genai_types.Type.STRING,
                    description="Which calendar the event is on.",
                    enum=["google", "outlook"],
                ),
            },
            required=["event_id", "source"],
        ),
    ),
    # -------- Event-driven triggers ------------------------------------------
    genai_types.FunctionDeclaration(
        name="create_email_trigger",
        description=(
            "Set up a condition-based trigger — NOT time-based like "
            "create_scheduled_task — that fires the next time a matching "
            "email arrives, e.g. 'notify me the moment I get an email from "
            "my boss' or 'let me know when an invoice email shows up'. "
            "Checked every minute. Requires sender_filter and/or "
            "keyword_filter — at least one. Confirm the exact condition, "
            "the prompt to run, and delivery channel with the user before "
            "creating, same as scheduled tasks."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "name": _schema(genai_types.Type.STRING, "Descriptive name for the trigger."),
                "source": genai_types.Schema(
                    type=genai_types.Type.STRING, enum=["gmail", "outlook"],
                    description="Which mailbox to watch.",
                ),
                "sender_filter": _schema(genai_types.Type.STRING, "Email address or domain to match the sender against (optional if keyword_filter given)."),
                "keyword_filter": _schema(genai_types.Type.STRING, "Word/phrase to match in subject or body (optional if sender_filter given)."),
                "prompt": _schema(genai_types.Type.STRING, "What Kin should do when a matching email arrives."),
                "channel": genai_types.Schema(
                    type=genai_types.Type.STRING, enum=["web", "email"],
                    description="Where to deliver the result.",
                ),
            },
            required=["name", "source", "prompt", "channel"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="list_email_triggers",
        description="List the user's active/inactive email-based triggers.",
        parameters=genai_types.Schema(type=genai_types.Type.OBJECT, properties={}),
    ),
    genai_types.FunctionDeclaration(
        name="delete_email_trigger",
        description="Delete an email trigger by its id (from list_email_triggers).",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={"trigger_id": _schema(genai_types.Type.STRING, "The trigger's id.")},
            required=["trigger_id"],
        ),
    ),
    # -------- Custom commands -----------------------------------------------
    genai_types.FunctionDeclaration(
        name="create_custom_command",
        description=(
            "Save a reusable prompt shortcut the user can trigger by typing "
            "'/name' in any future message — e.g. name='standup', "
            "prompt_template='Summarize today's calendar and top 3 tasks in "
            "3 bullet points.' Use when the user says things like 'save this "
            "as a shortcut' or 'make a /command for X'. Confirm the name and "
            "exact prompt with the user before saving if either was implied "
            "rather than explicitly stated."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "name": _schema(genai_types.Type.STRING, "Command name without the slash, e.g. 'standup'. Letters/numbers/underscores only."),
                "prompt_template": _schema(genai_types.Type.STRING, "The full prompt to run when this command is triggered."),
            },
            required=["name", "prompt_template"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="list_custom_commands",
        description="List the user's saved /command shortcuts.",
        parameters=genai_types.Schema(type=genai_types.Type.OBJECT, properties={}),
    ),
    genai_types.FunctionDeclaration(
        name="delete_custom_command",
        description="Delete a saved /command shortcut by name.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "name": _schema(genai_types.Type.STRING, "Command name without the slash."),
            },
            required=["name"],
        ),
    ),
    # -------- Expense logging -----------------------------------------------
    genai_types.FunctionDeclaration(
        name="find_receipt_emails",
        description=(
            "Gmail only. Find likely receipt/invoice/order-confirmation "
            "emails and return each with a body excerpt. Use this when the "
            "user asks to log expenses, track spending, or find receipts "
            "from their inbox. Read the body_excerpt of each result yourself "
            "to extract vendor, amount, currency, and date — this tool "
            "deliberately does NOT parse amounts itself (formats vary too "
            "much for a fixed parser); that's your job. After extracting, "
            "use create_spreadsheet (if the user doesn't have one for this "
            "yet) and update_sheet_values to log the rows — always show the "
            "user what you extracted before writing, since receipt parsing "
            "can be wrong and this is financial data."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "days": _schema(genai_types.Type.INTEGER, "How many days back to scan (default 30)."),
                "limit": _schema(genai_types.Type.INTEGER, "Max emails to check (default 20, max 30)."),
            },
        ),
    ),
    # -------- Inbox declutter ----------------------------------------------
    genai_types.FunctionDeclaration(
        name="list_promotional_senders",
        description=(
            "Gmail only. Find newsletters/promotional email cluttering the "
            "inbox, aggregated by sender with counts, so the user can decide "
            "what to clean up. Use this FIRST when the user asks to declutter, "
            "clean up, or unsubscribe from stuff in their inbox — never guess "
            "at senders, always show real counts from this tool."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "days": _schema(genai_types.Type.INTEGER, "How many days back to scan (default 30)."),
            },
        ),
    ),
    genai_types.FunctionDeclaration(
        name="declutter_gmail_sender",
        description=(
            "Gmail only. Bulk-archive or bulk-trash every promotional message "
            "from one sender at once. Always call list_promotional_senders "
            "first and confirm with the user which sender(s) and which action "
            "(archive vs. permanently trash) before calling this — it acts on "
            "every matching message in one shot, there's no per-message undo."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "sender_query": _schema(genai_types.Type.STRING, "Sender email address or domain, exactly as shown by list_promotional_senders."),
                "action": genai_types.Schema(
                    type=genai_types.Type.STRING,
                    description="'archive' (remove from inbox, recoverable) or 'trash' (Gmail trash, auto-deletes in 30 days).",
                    enum=["archive", "trash"],
                ),
                "days": _schema(genai_types.Type.INTEGER, "Same window used in list_promotional_senders (default 30)."),
            },
            required=["sender_query", "action"],
        ),
    ),
    # -------- Memory ------------------------------------------------------
    genai_types.FunctionDeclaration(
        name="remember_this",
        description=(
            "Save a new long-term memory RIGHT NOW. Call this whenever the user "
            "explicitly asks you to remember, note, or save something (e.g. "
            "'remember that...', 'note that...', 'keep in mind that...', "
            "'save this'). Don't wait for background extraction — this "
            "guarantees it's saved and lets you confirm it back to the user in "
            "the same reply. Do NOT call this for facts the user hasn't asked "
            "you to remember; ordinary conversation is still captured passively."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "content": _schema(
                    genai_types.Type.STRING,
                    "The fact to remember, written as a standalone statement (not 'remember that...').",
                ),
                "kind": genai_types.Schema(
                    type=genai_types.Type.STRING,
                    enum=sorted(mem.ALLOWED_KINDS),
                    description="preference | fact | event | relationship | goal | habit. Default fact.",
                ),
            },
            required=["content"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="recall_memory",
        description=(
            "Search long-term memory for facts you've previously noted about the "
            "user — preferences, relationships, ongoing goals, habits."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "query": _schema(genai_types.Type.STRING, "Natural-language description of what you're recalling."),
                "limit": _schema(genai_types.Type.INTEGER, "Max memories (default 5, max 10)."),
            },
            required=["query"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="forget_memory",
        description=(
            "Delete one specific long-term memory by its id. Use this when the "
            "user corrects something you remembered wrong, says it's outdated "
            "(they moved, changed jobs, a preference changed), or explicitly "
            "asks you to forget something. The memory's id is shown alongside "
            "it — either in the 'Things you remember' block already in this "
            "conversation, or in recall_memory's results if you need to look "
            "it up first. Never guess an id; only use one you've actually seen."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "memory_id": _schema(genai_types.Type.STRING, "The id of the memory to delete."),
            },
            required=["memory_id"],
        ),
    ),
    # -------- Scheduled Tasks ---------------------------------------------
    genai_types.FunctionDeclaration(
        name="create_scheduled_task",
        description=(
            "Schedule an automated natural language prompt routine to run on a cron schedule "
            "and deliver the results to Web, Telegram, or Email. The agent translates user instructions into cron and timezone."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "name": _schema(genai_types.Type.STRING, "Descriptive name for the scheduled task."),
                "prompt": _schema(genai_types.Type.STRING, "The prompt to run automatically (e.g., 'Summarize my day')."),
                "cron_expression": _schema(genai_types.Type.STRING, "A standard 5-field cron expression (e.g., '0 7 * * *' for 7 AM daily)."),
                "timezone": _schema(genai_types.Type.STRING, "The timezone to evaluate the cron in (e.g., 'America/New_York', 'UTC'). Defaults to UTC."),
                "channel": genai_types.Schema(
                    type=genai_types.Type.STRING,
                    description="Channel to send the result to. Must be 'web', or 'email'.",
                    enum=["web", "email"]
                ),
            },
            required=["name", "prompt", "cron_expression", "channel"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="list_scheduled_tasks",
        description="List all currently active and inactive scheduled tasks.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT, properties={}
        ),
    ),
    genai_types.FunctionDeclaration(
        name="delete_scheduled_task",
        description="Delete a scheduled task by its UUID.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "task_id": _schema(genai_types.Type.STRING, "UUID of the scheduled task to delete."),
            },
            required=["task_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="create_lead",
        description=(
            "Record visitor details (name, email, phone, company, job_title, country, industry, budget, etc.) as a business lead. "
            "Call this when the visitor shares their contact info, or after booking a meeting."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "bot_id": _schema(genai_types.Type.STRING, "UUID of the chatbot widget."),
                "name": _schema(genai_types.Type.STRING, "Visitor's full name."),
                "email": _schema(genai_types.Type.STRING, "Visitor's email address."),
                "phone": _schema(genai_types.Type.STRING, "Visitor's phone number."),
                "company": _schema(genai_types.Type.STRING, "Visitor's company name."),
                "job_title": _schema(genai_types.Type.STRING, "Visitor's job title."),
                "country": _schema(genai_types.Type.STRING, "Visitor's country."),
                "industry": _schema(genai_types.Type.STRING, "Visitor's industry."),
                "budget": _schema(genai_types.Type.STRING, "Visitor's budget."),
            },
            required=["bot_id", "name", "email"],
        ),
    ),
    # -------- Social scheduling (Postiz equivalent) -------------------------
    genai_types.FunctionDeclaration(
        name="list_connected_social_accounts",
        description=(
            "List which social media platforms (X, LinkedIn, Instagram, Facebook, "
            "YouTube, TikTok, Reddit, and 25+ others) the user has connected for "
            "post scheduling, and which are still available to connect. Call this "
            "before create_social_post if you're not sure the target platform is "
            "connected, or when the user asks 'what accounts do I have connected'."
        ),
        parameters=genai_types.Schema(type=genai_types.Type.OBJECT, properties={}),
    ),
    genai_types.FunctionDeclaration(
        name="create_social_post",
        description=(
            "Schedule (or immediately queue) a social media post on a platform the "
            "user has already connected — call list_connected_social_accounts first "
            "if you haven't already confirmed that in this conversation. "
            "publish_date should be ISO 8601 in the user's local timezone (e.g. "
            "'2026-05-12T09:00:00'); to post as soon as possible, use the current "
            "time. ALWAYS show the user the exact post content and platform and get "
            "their confirmation before calling this, unless they've already given "
            "you the final wording and clearly said to go ahead."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "integration_slug": _schema(
                    genai_types.Type.STRING,
                    "Platform slug, e.g. 'x', 'linkedin', 'instagram', 'facebook', 'youtube', 'tiktok', 'reddit'.",
                ),
                "account_id": _schema(
                    genai_types.Type.STRING,
                    "The specific connected account's id (from list_connected_social_accounts). Required if the "
                    "user has more than one connected account on the target platform — otherwise omit it and the "
                    "single connected account for that platform is used automatically.",
                ),
                "content": _schema(genai_types.Type.STRING, "The post text."),
                "publish_date": _schema(genai_types.Type.STRING, "ISO 8601 datetime in the user's local timezone."),
                "image_url": _schema(genai_types.Type.STRING, "Optional image/video URL to attach."),
                "draft": _schema(genai_types.Type.BOOLEAN, "True to save as a draft instead of queuing it for publish."),
            },
            required=["integration_slug", "content", "publish_date"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="list_social_posts",
        description="List the user's scheduled/draft/published/failed social posts, most recent first.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "state": _schema(genai_types.Type.STRING, "Filter: 'queue', 'draft', 'published', or 'failed'. Omit for all."),
                "limit": _schema(genai_types.Type.INTEGER, "Max results (default 10)."),
            },
        ),
    ),
    genai_types.FunctionDeclaration(
        name="cancel_social_post",
        description="Cancel/delete a queued or draft social post by its id (from list_social_posts). Cannot un-publish an already-published post.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={"post_id": _schema(genai_types.Type.STRING, "The post's id.")},
            required=["post_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="update_social_post",
        description=(
            "Change the content and/or publish_date of a social post you already scheduled (queue or draft state) "
            "by its post_id (from list_social_posts or a prior create_social_post result). Use this instead of "
            "create_social_post whenever the user is changing their mind about a post that already exists — "
            "calling create_social_post again for the same request creates a duplicate post."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "post_id": _schema(genai_types.Type.STRING, "The post's id."),
                "content": _schema(genai_types.Type.STRING, "New post text. Omit to leave unchanged."),
                "publish_date": _schema(genai_types.Type.STRING, "New ISO 8601 datetime in the user's local timezone. Omit to leave unchanged."),
                "draft": _schema(genai_types.Type.BOOLEAN, "Set true to move it to draft, false to move it to the publish queue. Omit to leave unchanged."),
            },
            required=["post_id"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="get_social_analytics",
        description="Get aggregate engagement analytics (impressions, likes, reposts, clicks) across all the user's published social posts.",
        parameters=genai_types.Schema(type=genai_types.Type.OBJECT, properties={}),
    ),
    genai_types.FunctionDeclaration(
        name="web_search",
        description=(
            "Search the live public web for current, real-world information — news, prices, sports scores, "
            "recent events, facts about people/companies/products, anything that might have changed since "
            "training or that you're not confident about. Returns a text summary of the top results with "
            "source URLs. Use this instead of guessing or saying you don't have access to the internet."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={"query": _schema(genai_types.Type.STRING, "The search query.")},
            required=["query"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="read_webpage",
        description=(
            "Fetch the full readable content of a specific URL (an article, docs page, product page, etc.) as "
            "plain text/markdown. Use this after web_search when a result's snippet isn't enough and you need "
            "the actual page content, or whenever the user shares a URL and asks about its content."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={"url": _schema(genai_types.Type.STRING, "The full URL to read, including https://.")},
            required=["url"],
        ),
    ),
    # -------- Voice Agents (LiveKit-powered phone agents) --------------------
    genai_types.FunctionDeclaration(
        name="create_voice_agent",
        description=(
            "Create a draft AI voice (phone call) agent for sales calls or receptionist duties, from a "
            "natural-language description of what the user wants. This only creates a DRAFT — no phone number is "
            "attached yet (provisioning a number costs money), and no calls will be made. After calling this, tell "
            "the user their agent was created and that they can review/tweak it and connect a phone number on the "
            "Voice Agents page of the dashboard. Infer sensible values for every field from the user's request; "
            "don't ask clarifying questions unless the request is too vague to name the agent or write a persona."
        ),
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "name": _schema(genai_types.Type.STRING, "Short human-readable name for the agent, e.g. 'Outbound Sales — Acme'."),
                "use_case": _schema(genai_types.Type.STRING, "One of: sales, receptionist, custom."),
                "persona": _schema(
                    genai_types.Type.STRING,
                    "The full system prompt / persona for the voice agent: who it is, the business it represents, "
                    "its goal on a call, tone, and any constraints. Write this out in full — it's what the agent "
                    "will actually be instructed with, not a summary.",
                ),
                "greeting": _schema(genai_types.Type.STRING, "Optional first line the agent opens the call with."),
                "tools": genai_types.Schema(
                    type=genai_types.Type.ARRAY,
                    items=genai_types.Schema(type=genai_types.Type.STRING),
                    description=(
                        "Which existing tool names to enable for this agent during calls, chosen from: "
                        "create_calendar_event, check_calendar_availability, create_lead. Include create_lead for "
                        "sales agents that should capture caller details, and calendar tools for anything that "
                        "books meetings/demos/appointments."
                    ),
                ),
            },
            required=["name", "use_case", "persona"],
        ),
    ),
    genai_types.FunctionDeclaration(
        name="list_voice_agents",
        description="List all the AI voice (phone) agents configured by the user.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={},
        ),
    ),
    genai_types.FunctionDeclaration(
        name="delete_voice_agent",
        description="Delete a voice agent using its unique ID.",
        parameters=genai_types.Schema(
            type=genai_types.Type.OBJECT,
            properties={
                "agent_id": _schema(genai_types.Type.STRING, "The unique ID of the voice agent to delete."),
            },
            required=["agent_id"],
        ),
    ),
]


def translate_schema(schema: dict) -> genai_types.Schema:
    if not isinstance(schema, dict):
        return genai_types.Schema(type=genai_types.Type.STRING)
    
    t_str = schema.get("type", "string").upper()
    try:
        t_enum = genai_types.Type[t_str]
    except KeyError:
        t_enum = genai_types.Type.STRING
        
    properties = {}
    if "properties" in schema and isinstance(schema["properties"], dict):
        for k, v in schema["properties"].items():
            properties[k] = translate_schema(v)
            
    items = None
    if "items" in schema:
        items = translate_schema(schema["items"])
        
    return genai_types.Schema(
        type=t_enum,
        description=schema.get("description", ""),
        properties=properties or None,
        required=schema.get("required") or None,
        items=items,
        enum=schema.get("enum") or None
    )


# ---------------------------------------------------------------------------
# Tool-subset routing — every DECLARATIONS entry used to be sent to the model
# on every single turn (133+ function schemas, even for "hi"). This groups
# tools into domains and only sends the domains a keyword match in the
# user's message suggests are relevant, falling back to the FULL set whenever
# nothing matches confidently — so a missed keyword costs nothing (same
# behavior as before), it only saves tokens on the turns where we're sure.
# ---------------------------------------------------------------------------

TOOL_CATEGORIES: dict[str, list[str]] = {
    "gmail": [
        "send_followup_nudge", "read_gmail", "get_gmail_message", "send_email",
        "reply_email", "draft_email", "modify_email_labels", "trash_email",
        "list_gmail_labels", "create_gmail_label", "delete_email_permanent",
        "list_drafts", "get_draft", "delete_draft", "list_email_threads",
        "get_email_thread", "reply_to_thread", "trash_email_thread",
        "modify_thread_labels", "find_receipt_emails", "list_promotional_senders",
        "declutter_gmail_sender", "create_email_trigger", "list_email_triggers",
        "delete_email_trigger",
    ],
    "calendar": [
        "read_calendar", "get_calendar_event", "create_calendar_event",
        "update_calendar_event", "delete_calendar_event", "check_calendar_availability",
        "prepare_meeting_brief",
    ],
    "tasks": [
        "read_tasks", "list_google_task_lists", "get_google_task", "list_google_tasks",
        "create_google_task", "update_google_task", "delete_google_task",
        "list_todo_lists", "list_todo_tasks", "create_todo_task", "update_todo_task",
        "delete_todo_task", "create_todo_list",
    ],
    "contacts": [
        "read_contacts", "list_google_contacts", "get_google_contact",
        "create_google_contact", "update_google_contact", "list_outlook_contacts",
        "get_outlook_contact", "create_outlook_contact", "delete_outlook_contact",
    ],
    "documents": [
        "search_documents", "read_full_document", "list_drive_files", "index_drive_file",
        "index_drive_folder", "index_onedrive_file", "index_onedrive_folder",
        "get_pdf_info", "get_spreadsheet_info", "send_file_to_user", "read_google_doc",
        "read_google_sheet", "list_sheet_tabs", "read_google_slides", "save_to_drive",
        "create_drive_folder", "copy_drive_file", "move_drive_file", "rename_drive_file",
        "delete_drive_item", "share_drive_item", "create_google_doc",
        "append_to_google_doc", "replace_text_in_doc", "replace_text_in_slides",
        "create_spreadsheet", "add_sheet_tab", "append_sheet_row", "update_sheet_values",
        "clear_sheet_range", "list_onedrive_files", "get_onedrive_metadata",
        "save_to_onedrive", "create_onedrive_folder", "rename_onedrive_item",
        "move_onedrive_item", "copy_onedrive_item", "delete_onedrive_item",
        "share_onedrive_item",
    ],
    "outlook": [
        "read_outlook", "get_outlook_message", "send_outlook_email", "reply_outlook_email",
        "mark_outlook_read", "delete_outlook_message", "move_outlook_message",
        "list_outlook_folders", "create_outlook_draft", "update_outlook_draft",
        "send_outlook_draft", "update_outlook_message", "list_outlook_folder_messages",
        "create_outlook_folder", "get_outlook_folder", "list_outlook_calendars",
        "create_outlook_calendar", "get_outlook_calendar", "list_outlook_events",
        "get_outlook_event", "create_outlook_event", "update_outlook_event",
        "delete_outlook_event", "list_outlook_attachments", "get_outlook_attachment",
    ],
    "scheduled_tasks": [
        "create_scheduled_task", "list_scheduled_tasks", "delete_scheduled_task",
        "create_custom_command", "list_custom_commands", "delete_custom_command",
    ],
    "social": [
        "list_connected_social_accounts", "create_social_post", "list_social_posts",
        "cancel_social_post", "update_social_post", "get_social_analytics",
    ],
    "leads": ["create_lead"],
    "voice_agents": ["create_voice_agent", "list_voice_agents", "delete_voice_agent"],
}

# Cheap and broadly useful enough (5 tools total) to just always include
# rather than try to keyword-match — memory recall in particular should
# never silently disappear because a message didn't happen to say "remember".
ALWAYS_INCLUDE_TOOLS = ["remember_this", "recall_memory", "forget_memory", "web_search", "read_webpage"]

CATEGORY_KEYWORDS: dict[str, list[str]] = {
    "gmail": ["email", "e-mail", "gmail", "inbox", "mail", "unsubscribe", "newsletter", "receipt"],
    "calendar": ["calendar", "meeting", "event", "schedule a", "appointment", "invite", "availability", "free/busy", "reschedule"],
    "tasks": ["task", "todo", "to-do", "to do list", "checklist"],
    "contacts": ["contact", "phone number", "address book"],
    "documents": ["document", "drive", "file", "pdf", "spreadsheet", "sheet", "slides", "onedrive", "folder", "doc "],
    "outlook": ["outlook", "microsoft 365", "office 365"],
    "scheduled_tasks": ["scheduled task", "/schedule", "recurring", "custom command", "automation"],
    "social": ["post to", "tweet", "linkedin", "instagram", "facebook", "social media", "twitter", "reddit", "tiktok", "youtube video", "schedule a post", "social post"],
    "leads": ["lead", "prospect"],
    "voice_agents": ["voice agent", "phone agent", "sales agent", "sales call", "receptionist", "voice assistant", "phone call agent", "calling agent", "cold call", "answer calls", "answer the phone"],
}


def select_relevant_tool_names(text: str) -> Optional[set[str]]:
    """Returns the union of tool names for every category whose keywords
    appear in `text`, plus ALWAYS_INCLUDE_TOOLS — or None if nothing matched,
    which the caller should treat as "include everything" (safe fallback,
    identical to the old always-send-all-tools behavior)."""
    lower = (text or "").lower()
    if not lower:
        return None
    matched: set[str] = set()
    for category, keywords in CATEGORY_KEYWORDS.items():
        if any(kw in lower for kw in keywords):
            matched.update(TOOL_CATEGORIES[category])
    if not matched:
        return None
    matched.update(ALWAYS_INCLUDE_TOOLS)
    return matched


def get_tool_config(
    mcp_servers: list,
    exclude_tools: list[str] = None,
    user_id: str = None,
    supabase = None,
    message_text: str = "",
) -> genai_types.Tool:
    if exclude_tools is None:
        exclude_tools = []
    relevant = select_relevant_tool_names(message_text)
    if relevant is not None:
        decls = [d for d in DECLARATIONS if d.name not in exclude_tools and d.name in relevant]
    else:
        decls = [d for d in DECLARATIONS if d.name not in exclude_tools]

    # 1. Fetch remote MCP server tools
    for server in mcp_servers:
        server_name = server.get("name", "")
        tools_list = server.get("tools") or []
        for tool in tools_list:
            prefixed_name = f"{server_name}__{tool['name']}"
            if prefixed_name in exclude_tools:
                continue
            input_schema = tool.get("inputSchema", {})
            params = translate_schema(input_schema)
            decl = genai_types.FunctionDeclaration(
                name=prefixed_name,
                description=tool.get("description", ""),
                parameters=params
            )
            decls.append(decl)
            
    # 2. Fetch custom integrations installed by this user
    if user_id and supabase:
        try:
            res = supabase.table("integration_installs").select("integration_slug, integrations(name, manifest)").eq("user_id", user_id).execute()
            for install in res.data or []:
                slug = install["integration_slug"]
                integ = install.get("integrations") or {}
                manifest = integ.get("manifest") or {}
                actions = manifest.get("actions") or []
                for action in actions:
                    action_name = action.get("name")
                    prefixed_name = f"integration__{slug}__{action_name}"
                    if prefixed_name in exclude_tools:
                        continue
                    
                    properties = {}
                    required = []
                    for inp in action.get("inputs") or []:
                        inp_name = inp.get("name")
                        inp_type = inp.get("type", "string")
                        inp_desc = inp.get("description", "")
                        properties[inp_name] = {
                            "type": inp_type if inp_type in ["string", "number", "integer", "boolean", "array", "object"] else "string",
                            "description": inp_desc
                        }
                        if inp.get("required"):
                            required.append(inp_name)
                    
                    decl = genai_types.FunctionDeclaration(
                        name=prefixed_name,
                        description=action.get("description", ""),
                        parameters=genai_types.Schema(
                            type=genai_types.Type.OBJECT,
                            properties=properties or None,
                            required=required or None
                        )
                    )
                    decls.append(decl)
        except Exception:
            logger.exception("Failed to load custom integration declarations")
            
    return genai_types.Tool(function_declarations=decls)


def tool_config() -> genai_types.Tool:
    return genai_types.Tool(function_declarations=DECLARATIONS)


# ---------------------------------------------------------------------------
# Stage 4 (litellm migration): OpenAI/litellm-shaped tool declarations.
#
# DECLARATIONS above (and everything built from genai_types.Schema/Type in
# this file — translate_schema(), the custom-integration Schema construction
# in get_tool_config()) stays exactly as-is; nothing here mutates or
# replaces it. This section is a *mechanical, structural* conversion pass:
# it walks the existing genai_types.FunctionDeclaration/Schema objects (via
# pydantic's model_dump()) rather than hand-transcribing each of the ~142
# entries, so DECLARATIONS remains the single source of truth for tool
# schemas and OPENAI_DECLARATIONS is always derived from it, never
# hand-duplicated.
#
# Only agent_tools.execute()'s CALLERS need to change (which tool-schema
# list they hand to the model); execute()'s own dispatch contract
# (`execute(name, args, *, user, supabase, genai_client, context)`) is
# untouched — args always arrive as a plain dict regardless of whether the
# model called via Gemini-native function_call.args or litellm/OpenAI
# tool_calls[].function.arguments (app/core/llm.py's parsed_tool_calls()
# JSON-decodes that side back into a dict before it ever reaches execute()).
# ---------------------------------------------------------------------------


def _lower_schema_types(node: Any) -> Any:
    """Recursively rewrite genai_types.Type enum members (e.g. Type.STRING)
    into lowercase JSON-schema type strings ("string") anywhere they appear
    in a dict produced by Schema.model_dump() — including nested
    properties/items, which model_dump() only partially normalizes (nested
    Schema objects come back as dicts, but their `type` field stays an enum
    member, not a plain value)."""
    if isinstance(node, dict):
        out: dict[str, Any] = {}
        for key, value in node.items():
            if key == "type":
                out[key] = str(getattr(value, "value", value)).lower()
            elif key == "properties" and isinstance(value, dict):
                out[key] = {k: _lower_schema_types(v) for k, v in value.items()}
            elif key == "items":
                out[key] = _lower_schema_types(value)
            else:
                out[key] = value
        return out
    return node


def _schema_to_json_schema(schema: Optional[genai_types.Schema]) -> dict:
    """genai_types.Schema -> plain OpenAI/JSON-Schema dict."""
    if schema is None:
        return {"type": "object", "properties": {}}
    raw = schema.model_dump(exclude_none=True)
    normalized = _lower_schema_types(raw)
    normalized.setdefault("type", "object")
    if normalized["type"] == "object":
        normalized.setdefault("properties", {})
    return normalized


def _declaration_to_openai_tool(decl: genai_types.FunctionDeclaration) -> dict:
    """genai_types.FunctionDeclaration -> OpenAI/litellm tool dict:
    ``{"type": "function", "function": {"name", "description", "parameters"}}``.
    """
    return {
        "type": "function",
        "function": {
            "name": decl.name,
            "description": decl.description or "",
            "parameters": _schema_to_json_schema(decl.parameters),
        },
    }


# Derived, not hand-written: one entry per DECLARATIONS entry, same order,
# same names — see the module-level comment above.
OPENAI_DECLARATIONS: list[dict] = [_declaration_to_openai_tool(d) for d in DECLARATIONS]


def get_openai_tool_config(
    mcp_servers: list,
    exclude_tools: list[str] = None,
    user_id: str = None,
    supabase=None,
    message_text: str = "",
) -> list[dict]:
    """litellm/OpenAI-shaped counterpart to get_tool_config(): same
    filtering (relevance-based tool subset via select_relevant_tool_names(),
    exclude_tools, per-user MCP servers, per-user custom integrations), but
    returns a flat ``list[dict]`` of OpenAI tool defs (litellm's expected
    ``tools=`` shape) instead of a single Gemini ``genai_types.Tool``.
    """
    if exclude_tools is None:
        exclude_tools = []
    relevant = select_relevant_tool_names(message_text)
    if relevant is not None:
        decls = [d for d in DECLARATIONS if d.name not in exclude_tools and d.name in relevant]
    else:
        decls = [d for d in DECLARATIONS if d.name not in exclude_tools]
    tools: list[dict] = [_declaration_to_openai_tool(d) for d in decls]

    # 1. Remote MCP server tools — same translate_schema() Gemini-Schema
    # construction as get_tool_config(), then converted the same way as
    # DECLARATIONS above (translate_schema's output is a genai_types.Schema,
    # so it goes through the identical _schema_to_json_schema() path).
    for server in mcp_servers:
        server_name = server.get("name", "")
        tools_list = server.get("tools") or []
        for tool in tools_list:
            prefixed_name = f"{server_name}__{tool['name']}"
            if prefixed_name in exclude_tools:
                continue
            input_schema = tool.get("inputSchema", {})
            params = translate_schema(input_schema)
            tools.append({
                "type": "function",
                "function": {
                    "name": prefixed_name,
                    "description": tool.get("description", ""),
                    "parameters": _schema_to_json_schema(params),
                },
            })

    # 2. Per-user custom integrations — these are already built from plain
    # OpenAI-style {"type", "description"} dicts in get_tool_config(), so no
    # Gemini-Schema round trip is needed; just assemble the OpenAI tool dict
    # directly.
    if user_id and supabase:
        try:
            res = (
                supabase.table("integration_installs")
                .select("integration_slug, integrations(name, manifest)")
                .eq("user_id", user_id)
                .execute()
            )
            for install in res.data or []:
                slug = install["integration_slug"]
                integ = install.get("integrations") or {}
                manifest = integ.get("manifest") or {}
                actions = manifest.get("actions") or []
                for action in actions:
                    action_name = action.get("name")
                    prefixed_name = f"integration__{slug}__{action_name}"
                    if prefixed_name in exclude_tools:
                        continue

                    properties = {}
                    required = []
                    for inp in action.get("inputs") or []:
                        inp_name = inp.get("name")
                        inp_type = inp.get("type", "string")
                        inp_desc = inp.get("description", "")
                        properties[inp_name] = {
                            "type": inp_type if inp_type in ["string", "number", "integer", "boolean", "array", "object"] else "string",
                            "description": inp_desc,
                        }
                        if inp.get("required"):
                            required.append(inp_name)

                    tools.append({
                        "type": "function",
                        "function": {
                            "name": prefixed_name,
                            "description": action.get("description", ""),
                            "parameters": {
                                "type": "object",
                                "properties": properties,
                                "required": required,
                            },
                        },
                    })
        except Exception:
            logger.exception("Failed to load custom integration declarations (openai path)")

    return tools


# ---------------------------------------------------------------------------
# Handlers
# ---------------------------------------------------------------------------


def _need_google(user: dict[str, Any]) -> dict[str, Any] | None:
    if not user.get("google_access_token"):
        return {
            "error": "Google not connected. Ask the user to visit /dashboard/integrations to connect their Google account.",
        }
    return None


def _need_microsoft(
    user: dict[str, Any], *, required_scope: Optional[str] = None
) -> dict[str, Any] | None:
    """Guard a Microsoft tool call.

    Returns None if everything's good, or a dict with an `error` describing
    what the user must do.  When `required_scope` is given (e.g.
    "Calendars.ReadWrite"), checks the granted scope set stored at connect
    time. If the scope wasn't granted, Graph would return either 403 or
    a silent empty result — neither is useful to the user. Surface the real
    fix instead: "reconnect Microsoft."
    """
    if not user.get("microsoft_access_token"):
        return {
            "error": "Microsoft not connected. Ask the user to visit /dashboard/integrations to connect their Microsoft 365 account.",
        }
    if required_scope:
        granted = (user.get("microsoft_scopes") or "").lower()
        # Graph returns scopes space-separated (or comma in some flows).
        if required_scope.lower() not in granted:
            return {
                "error": (
                    f"Microsoft is connected but the '{required_scope}' "
                    "permission was not granted. Tell the user: 'Your "
                    "Microsoft connection is missing calendar/contacts "
                    "permission. Please go to /dashboard/integrations, "
                    "click Disconnect Microsoft, then Connect Microsoft "
                    "again — the new consent screen will ask for the "
                    "calendar and contacts permissions.'"
                ),
            }
    return None


def _parse_iso(s: str) -> datetime:
    # Accept ISO with offset, "...Z", space separator, or date-only.
    s = (s or "").strip()
    if not s:
        raise ValueError("empty datetime")
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    for candidate in (s, s.replace(" ", "T"), s + "T00:00:00"):
        try:
            return datetime.fromisoformat(candidate)
        except ValueError:
            continue
    raise ValueError(f"unparseable datetime: {s!r}")


# ---- Gmail ---------------------------------------------------------------


async def _extra_google_accounts(supabase, user_id: str) -> list[dict]:
    """Pro/Executive users can connect up to 2 extra Google accounts (see
    main.py's /api/integrations/google/accounts) on top of their primary
    one. Read-only aggregation lives here; empty for everyone who hasn't
    connected any, so this is a no-op for the vast majority of users."""
    try:
        res = (
            supabase.table("kin_connected_accounts")
            .select("*")
            .eq("user_id", user_id)
            .execute()
        )
        return res.data or []
    except Exception:  # noqa: BLE001
        logger.exception("failed to load extra connected accounts for %s", user_id)
        return []


async def _read_gmail(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    raw_query = (args.get("query") or "").strip()
    # When the user provides an explicit search query (date range, sender, etc.),
    # do NOT append the default category filter — it hides emails in promotions/
    # social tabs that the user explicitly asked about.  Only apply the filter
    # for the "show me my recent inbox" default case.
    if raw_query:
        query = raw_query
    else:
        query = "-category:promotions -category:social"
    limit = min(int(args.get("limit") or 10), 50)
    messages = await g.list_gmail_messages(supabase, user, limit=limit, query=query)
    for m in messages:
        m["account"] = user.get("google_email") or "primary"

    extras = await _extra_google_accounts(supabase, user["id"])
    for extra in extras:
        try:
            extra_messages = await g.list_gmail_messages(
                supabase, extra, limit=limit, query=query, table="kin_connected_accounts"
            )
        except Exception:  # noqa: BLE001
            logger.exception("gmail fetch failed for extra account %s", extra.get("id"))
            continue
        for m in extra_messages:
            m["account"] = extra.get("label") or extra.get("google_email") or "connected account"
        messages.extend(extra_messages)

    return {"count": len(messages), "messages": messages}


async def _get_gmail_message(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    mid = (args.get("message_id") or "").strip()
    if not mid:
        return {"error": "message_id required"}
    return await g.get_gmail_message(supabase, user, mid)


async def _send_followup_nudge(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    plan = (user.get("plan") or "free").lower()
    if plan not in _EMAIL_FOLLOWUP_PLANS:
        return {"error": "Email follow-ups are a Basic+ feature. Upgrade at /dashboard/billing."}
    ref = (args.get("nudge_ref") or "").strip().lower()
    if not ref:
        return {"error": "nudge_ref required"}
    res = (
        supabase.table("kin_email_watches")
        .select("*")
        .eq("user_id", user["id"])
        .is_("nudge_sent_at", "null")
        .is_("replied_at", "null")
        .eq("dismissed", False)
        .execute()
    )
    watch = next((w for w in (res.data or []) if str(w["id"]).replace("-", "").startswith(ref)), None)
    if not watch:
        return {"error": "Couldn't find that follow-up — it may already be handled or too old."}

    note = (args.get("note") or "").strip()
    body = "Hi — just checking in on this, let me know if you had a chance to look."
    if note:
        body += f"\n\n{note}"
    try:
        result = await g.reply_gmail_thread(supabase, user, thread_id=watch["thread_id"], body=body)
    except Exception as exc:  # noqa: BLE001
        logger.exception("failed to send follow-up nudge for watch %s", watch["id"])
        return {"error": f"Could not send the nudge: {exc}"}

    supabase.table("kin_email_watches").update({
        "nudge_sent_at": datetime.now(tz=timezone.utc).isoformat(),
    }).eq("id", watch["id"]).execute()
    return {"status": "sent", **result}


_EMAIL_FOLLOWUP_PLANS = {"basic", "pro", "executive"}  # kept in sync with main.py's PAID_PLANS


async def _watch_email_thread(user: dict, supabase, result: dict) -> None:
    """Track a Kin-sent thread so /cron/check-email-followups can notice a
    reply or offer to nudge a contact who's gone quiet. Basic+ feature —
    re-checked here (not just at the settings-write gate) in case the plan
    lapsed after the setting was turned on. Opt-out via Settings
    (email_followups_enabled); never blocks the send on failure."""
    plan = (user.get("plan") or "free").lower()
    if plan not in _EMAIL_FOLLOWUP_PLANS or not user.get("email_followups_enabled", True):
        return
    thread_id, message_id = result.get("thread_id"), result.get("id")
    if not thread_id or not message_id:
        return
    try:
        supabase.table("kin_email_watches").insert({
            "user_id": user["id"], "thread_id": thread_id, "sent_message_id": message_id,
        }).execute()
    except Exception:  # noqa: BLE001
        logger.exception("failed to create email watch for thread %s", thread_id)


def _render_signature(user: dict) -> str:
    """Plain-text email signature from the user's structured settings — '' if
    disabled or empty. Rendered server-side (not stored as raw HTML) so the
    dashboard preview and the actually-sent signature can never drift apart."""
    if not user.get("email_signature_enabled"):
        return ""
    lines: list[str] = []
    name = (user.get("email_signature_name") or "").strip()
    title = (user.get("email_signature_title") or "").strip()
    phone = (user.get("email_signature_phone") or "").strip()
    links = user.get("email_signature_links") or []
    if name:
        lines.append(name)
    if title:
        lines.append(title)
    if phone:
        lines.append(phone)
    for link in links:
        if not isinstance(link, dict):
            continue
        label = (link.get("label") or "").strip()
        url = (link.get("url") or "").strip()
        if not url:
            continue
        lines.append(f"{label}: {url}" if label else url)
    if not lines:
        return ""
    return "\n\n-- \n" + "\n".join(lines)


def _with_signature(body: str, user: dict) -> str:
    return (body or "") + _render_signature(user)


async def _resolve_attachment(args: dict, user: dict, supabase) -> tuple[Optional[list[dict]], Optional[dict]]:
    """(attachments, error). error is a dict to return as-is from the tool
    if the requested file couldn't be found/retrieved; attachments is a
    ready-to-send list for google_integrations' send/reply functions."""
    name = (args.get("attachment_file_name") or "").strip()
    if not name:
        return None, None
    doc = await doc_rag.get_document_bytes(supabase, user, name)
    if doc.get("error"):
        return None, {"error": doc["error"]}
    return [{
        "filename": doc["file_name"],
        "mime_type": doc.get("mime_type") or "application/octet-stream",
        "data": doc["data"],
    }], None


async def _send_email(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    to = args.get("to") or []
    if not to:
        return {"error": "at least one recipient required"}
    attachments, att_err = await _resolve_attachment(args, user, supabase)
    if att_err:
        return att_err
    result = await g.send_gmail(
        supabase,
        user,
        to=list(to),
        subject=args.get("subject") or "",
        body=_with_signature(args.get("body") or "", user),
        cc=list(args.get("cc") or []),
        bcc=list(args.get("bcc") or []),
        attachments=attachments,
    )
    await _watch_email_thread(user, supabase, result)
    return result


async def _reply_email(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    attachments, att_err = await _resolve_attachment(args, user, supabase)
    if att_err:
        return att_err
    result = await g.reply_gmail(
        supabase,
        user,
        message_id=args.get("message_id") or "",
        body=_with_signature(args.get("body") or "", user),
        attachments=attachments,
    )
    await _watch_email_thread(user, supabase, result)
    return result


async def _draft_email(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    to = args.get("to") or []
    return await g.create_gmail_draft(
        supabase,
        user,
        to=list(to),
        subject=args.get("subject") or "",
        body=args.get("body") or "",
        cc=list(args.get("cc") or []),
    )


async def _modify_email_labels(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    return await g.modify_gmail_labels(
        supabase,
        user,
        message_id=args.get("message_id") or "",
        add=list(args.get("add") or []),
        remove=list(args.get("remove") or []),
    )


async def _trash_email(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    return await g.trash_gmail(
        supabase, user, message_id=args.get("message_id") or ""
    )


async def _list_gmail_labels(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    labels = await g.list_gmail_labels(supabase, user)
    return {"count": len(labels), "labels": labels}


async def _create_gmail_label(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    name = (args.get("name") or "").strip()
    if not name:
        return {"error": "name required"}
    return await g.create_gmail_label(supabase, user, name=name)


async def _delete_email_permanent(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    return await g.delete_gmail_message(
        supabase, user, message_id=args.get("message_id") or ""
    )


# ---- Gmail drafts --------------------------------------------------------


async def _list_drafts(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    drafts = await g.list_gmail_drafts(
        supabase, user, limit=int(args.get("limit") or 20)
    )
    return {"count": len(drafts), "drafts": drafts}


async def _get_draft(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    return await g.get_gmail_draft(
        supabase, user, draft_id=args.get("draft_id") or ""
    )


async def _delete_draft(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    return await g.delete_gmail_draft(
        supabase, user, draft_id=args.get("draft_id") or ""
    )


# ---- Gmail threads -------------------------------------------------------


async def _list_email_threads(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    threads = await g.list_gmail_threads(
        supabase,
        user,
        query=args.get("query") or "",
        limit=int(args.get("limit") or 20),
    )
    return {"count": len(threads), "threads": threads}


async def _get_email_thread(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    return await g.get_gmail_thread(
        supabase, user, thread_id=args.get("thread_id") or ""
    )


async def _reply_to_thread(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    attachments, att_err = await _resolve_attachment(args, user, supabase)
    if att_err:
        return att_err
    result = await g.reply_gmail_thread(
        supabase,
        user,
        thread_id=args.get("thread_id") or "",
        body=_with_signature(args.get("body") or "", user),
        attachments=attachments,
    )
    await _watch_email_thread(user, supabase, result)
    return result


async def _trash_email_thread(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    return await g.trash_gmail_thread(
        supabase, user, thread_id=args.get("thread_id") or ""
    )


async def _modify_thread_labels(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    return await g.modify_gmail_thread_labels(
        supabase,
        user,
        thread_id=args.get("thread_id") or "",
        add=list(args.get("add") or []),
        remove=list(args.get("remove") or []),
    )


# ---- Single Google Task -------------------------------------------------


async def _get_google_task(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    return await g.get_google_task(
        supabase,
        user,
        task_id=args.get("task_id") or "",
        task_list_id=args.get("task_list_id") or "@default",
    )


# ---- Calendar ------------------------------------------------------------


async def _read_calendar(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    days = max(1, min(int(args.get("days_ahead") or 1), 30))
    now = datetime.now(tz=timezone.utc)
    time_min, time_max = now - timedelta(hours=2), now + timedelta(days=days)
    events = await g.list_calendar_events(supabase, user, time_min, time_max)
    for e in events:
        e["account"] = user.get("google_email") or "primary"

    extras = await _extra_google_accounts(supabase, user["id"])
    for extra in extras:
        try:
            extra_events = await g.list_calendar_events(
                supabase, extra, time_min, time_max, table="kin_connected_accounts"
            )
        except Exception:  # noqa: BLE001
            logger.exception("calendar fetch failed for extra account %s", extra.get("id"))
            continue
        for e in extra_events:
            e["account"] = extra.get("label") or extra.get("google_email") or "connected account"
        events.extend(extra_events)

    return {"count": len(events), "events": events}


async def _get_calendar_event(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    return await g.get_calendar_event(supabase, user, args.get("event_id") or "")


async def _create_calendar_event(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    return await g.create_calendar_event(
        supabase,
        user,
        summary=args.get("summary") or "",
        start=args.get("start") or "",
        end=args.get("end") or "",
        description=args.get("description"),
        location=args.get("location"),
        attendees=list(args.get("attendees") or []) or None,
        all_day=bool(args.get("all_day")),
    )


async def _update_calendar_event(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    return await g.update_calendar_event(
        supabase,
        user,
        event_id=args.get("event_id") or "",
        summary=args.get("summary"),
        start=args.get("start"),
        end=args.get("end"),
        description=args.get("description"),
        location=args.get("location"),
        attendees=(list(args["attendees"]) if "attendees" in args else None),
    )


async def _delete_calendar_event(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    return await g.delete_calendar_event(
        supabase, user, event_id=args.get("event_id") or ""
    )


async def _check_calendar_availability(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    try:
        time_min = _parse_iso(args.get("start") or "")
        time_max = _parse_iso(args.get("end") or "")
    except ValueError:
        return {"error": (
            "Invalid time format. Call this again with 'start' and 'end' as ISO 8601 "
            "datetimes INCLUDING the timezone offset, e.g. '2026-06-25T09:00:00+05:30'."
        )}
    if time_max <= time_min:
        return {"error": "'end' must be after 'start' — use a 30-minute window."}
    try:
        return await g.check_calendar_availability(
            supabase, user, time_min=time_min, time_max=time_max,
        )
    except g.GoogleNotConnected:
        raise
    except Exception as exc:  # noqa: BLE001
        logger.exception("check_calendar_availability failed")
        return {"error": f"availability check failed: {exc}"}


# ---- Google Tasks --------------------------------------------------------


async def _list_google_task_lists(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    return {"lists": await g.list_google_task_lists(supabase, user)}


async def _list_google_tasks(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    return {
        "tasks": await g.list_google_tasks(
            supabase,
            user,
            task_list_id=args.get("task_list_id") or "@default",
            show_completed=bool(args.get("show_completed")),
            limit=int(args.get("limit") or 50),
        )
    }


async def _create_google_task(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    return await g.create_google_task(
        supabase,
        user,
        title=args.get("title") or "",
        notes=args.get("notes"),
        due=args.get("due"),
        task_list_id=args.get("task_list_id") or "@default",
    )


async def _update_google_task(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    return await g.update_google_task(
        supabase,
        user,
        task_id=args.get("task_id") or "",
        title=args.get("title"),
        notes=args.get("notes"),
        due=args.get("due"),
        completed=(args["completed"] if "completed" in args else None),
        task_list_id=args.get("task_list_id") or "@default",
    )


async def _delete_google_task(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    return await g.delete_google_task(
        supabase,
        user,
        task_id=args.get("task_id") or "",
        task_list_id=args.get("task_list_id") or "@default",
    )


# ---- Google Contacts -----------------------------------------------------


async def _list_google_contacts(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    return {
        "contacts": await g.list_google_contacts(
            supabase,
            user,
            limit=int(args.get("limit") or 30),
            query=args.get("query") or "",
        )
    }


async def _get_google_contact(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    return await g.get_google_contact(
        supabase, user, resource_name=args.get("resource_name") or ""
    )


async def _create_google_contact(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    return await g.create_google_contact(
        supabase,
        user,
        name=args.get("name") or "",
        email=args.get("email"),
        phone=args.get("phone"),
        company=args.get("company"),
        notes=args.get("notes"),
    )


async def _update_google_contact(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    return await g.update_google_contact(
        supabase,
        user,
        resource_name=args.get("resource_name") or "",
        name=args.get("name"),
        email=args.get("email"),
        phone=args.get("phone"),
        company=args.get("company"),
        notes=args.get("notes"),
    )


# ---- Kin-local Supabase --------------------------------------------------


def _read_tasks(args: dict, user: dict, supabase) -> dict:
    status = (args.get("status") or "open").lower()
    limit = min(int(args.get("limit") or 20), 50)
    q = supabase.table("tasks").select(
        "title, description, status, priority, due_date, created_at"
    ).eq("user_id", user["id"])
    if status == "open":
        q = q.neq("status", "done")
    elif status in ("todo", "in_progress", "done"):
        q = q.eq("status", status)
    res = q.order("due_date", desc=False).limit(limit).execute()
    rows = res.data or []
    return {"count": len(rows), "tasks": rows}


def _read_contacts(args: dict, user: dict, supabase) -> dict:
    query = (args.get("query") or "").strip()
    limit = min(int(args.get("limit") or 10), 50)
    q = supabase.table("contacts").select(
        "name, email, phone, company, notes"
    ).eq("user_id", user["id"])
    if query:
        like = f"%{query}%"
        q = q.or_(
            f"name.ilike.{like},email.ilike.{like},company.ilike.{like}"
        )
    res = q.limit(limit).execute()
    rows = res.data or []
    return {"count": len(rows), "contacts": rows}


# ---- Microsoft Outlook ---------------------------------------------------


async def _read_outlook(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    msgs = await ms.list_outlook_messages(
        supabase,
        user,
        limit=int(args.get("limit") or 15),
        query=args.get("query") or "",
        folder=args.get("folder") or "Inbox",
        max_days_old=int(args.get("max_days_old")) if args.get("max_days_old") else None,
    )
    return {"count": len(msgs), "messages": msgs}


async def _get_outlook_message(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.get_outlook_message(
        supabase, user, message_id=args.get("message_id") or ""
    )


async def _send_outlook_email(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.send_outlook_message(
        supabase,
        user,
        to=list(args.get("to") or []),
        subject=args.get("subject") or "",
        body=args.get("body") or "",
        cc=list(args.get("cc") or []) or None,
        bcc=list(args.get("bcc") or []) or None,
    )


async def _reply_outlook_email(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.reply_outlook_message(
        supabase,
        user,
        message_id=args.get("message_id") or "",
        body=args.get("body") or "",
    )


async def _mark_outlook_read(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.mark_outlook_read(
        supabase,
        user,
        message_id=args.get("message_id") or "",
        is_read=bool(args.get("is_read", True)),
    )


async def _delete_outlook_message(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.delete_outlook_message(
        supabase, user, message_id=args.get("message_id") or ""
    )


async def _move_outlook_message(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.move_outlook_message(
        supabase,
        user,
        message_id=args.get("message_id") or "",
        destination_folder=args.get("destination_folder") or "archive",
    )


async def _list_outlook_folders(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return {"folders": await ms.list_outlook_folders(supabase, user)}


# ---- Microsoft OneDrive --------------------------------------------------


async def _list_onedrive_files(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.list_onedrive_files(
        supabase,
        user,
        query=args.get("query") or None,
        folder_id=args.get("folder_id") or None,
        page_size=int(args.get("limit") or 30),
    )


async def _get_onedrive_metadata(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.get_onedrive_metadata(
        supabase, user, item_id=args.get("item_id") or ""
    )


# ---- Microsoft ToDo ------------------------------------------------------


async def _list_todo_lists(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return {"lists": await ms.list_todo_lists(supabase, user)}


async def _list_todo_tasks(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return {
        "tasks": await ms.list_todo_tasks(
            supabase,
            user,
            list_id=args.get("list_id"),
            show_completed=bool(args.get("show_completed")),
            limit=int(args.get("limit") or 50),
        )
    }


async def _create_todo_task(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.create_todo_task(
        supabase,
        user,
        title=args.get("title") or "",
        body=args.get("body"),
        due=args.get("due"),
        importance=args.get("importance") or "normal",
        list_id=args.get("list_id"),
    )


async def _update_todo_task(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.update_todo_task(
        supabase,
        user,
        task_id=args.get("task_id") or "",
        list_id=args.get("list_id") or "",
        title=args.get("title"),
        body=args.get("body"),
        completed=args["completed"] if "completed" in args else None,
        due=args.get("due"),
        importance=args.get("importance"),
    )


async def _delete_todo_task(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.delete_todo_task(
        supabase,
        user,
        task_id=args.get("task_id") or "",
        list_id=args.get("list_id") or "",
    )


# ---- OneDrive writes -----------------------------------------------------


async def _save_to_onedrive(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.upload_onedrive_text_file(
        supabase,
        user,
        name=args.get("name") or "Untitled.txt",
        text=args.get("text") or "",
        mime_type=args.get("mime_type") or "text/plain",
        parent_folder_id=args.get("parent_folder_id"),
    )


async def _create_onedrive_folder(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.create_onedrive_folder(
        supabase,
        user,
        name=args.get("name") or "",
        parent_folder_id=args.get("parent_folder_id"),
    )


async def _rename_onedrive_item(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.rename_onedrive_item(
        supabase,
        user,
        item_id=args.get("item_id") or "",
        new_name=args.get("new_name") or "",
    )


async def _move_onedrive_item(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.move_onedrive_item(
        supabase,
        user,
        item_id=args.get("item_id") or "",
        new_parent_folder_id=args.get("new_parent_folder_id") or "",
    )


async def _copy_onedrive_item(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.copy_onedrive_item(
        supabase,
        user,
        item_id=args.get("item_id") or "",
        new_name=args.get("new_name"),
        parent_folder_id=args.get("parent_folder_id"),
    )


async def _delete_onedrive_item(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.delete_onedrive_item(
        supabase, user, item_id=args.get("item_id") or ""
    )


async def _share_onedrive_item(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.share_onedrive_item(
        supabase,
        user,
        item_id=args.get("item_id") or "",
        email=args.get("email"),
        role=args.get("role") or "read",
    )


# ---- Outlook drafts ------------------------------------------------------


async def _create_outlook_draft(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.create_outlook_draft(
        supabase,
        user,
        to=list(args.get("to") or []),
        subject=args.get("subject") or "",
        body=args.get("body") or "",
        cc=list(args.get("cc") or []) or None,
    )


async def _update_outlook_draft(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.update_outlook_draft(
        supabase,
        user,
        draft_id=args.get("draft_id") or "",
        subject=args.get("subject"),
        body=args.get("body"),
        to=list(args["to"]) if "to" in args else None,
    )


async def _send_outlook_draft(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.send_outlook_draft(
        supabase, user, draft_id=args.get("draft_id") or ""
    )


async def _update_outlook_message(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.update_outlook_message(
        supabase,
        user,
        message_id=args.get("message_id") or "",
        is_read=args["is_read"] if "is_read" in args else None,
        flag=args.get("flag"),
        categories=list(args["categories"]) if "categories" in args else None,
    )


async def _list_outlook_folder_messages(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    msgs = await ms.list_outlook_folder_messages(
        supabase,
        user,
        folder_id=args.get("folder_id") or "",
        limit=int(args.get("limit") or 25),
    )
    return {"count": len(msgs), "messages": msgs}


async def _create_outlook_folder(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.create_outlook_folder(
        supabase,
        user,
        name=args.get("name") or "",
        parent_folder_id=args.get("parent_folder_id"),
    )


async def _get_outlook_folder(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.get_outlook_folder(
        supabase, user, folder_id=args.get("folder_id") or ""
    )


# ---- Outlook calendars + events ------------------------------------------


async def _list_outlook_calendars(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user, required_scope="Calendars.ReadWrite"):
        return g_err
    return {"calendars": await ms.list_outlook_calendars(supabase, user)}


async def _create_outlook_calendar(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user, required_scope="Calendars.ReadWrite"):
        return g_err
    return await ms.create_outlook_calendar(
        supabase, user, name=args.get("name") or ""
    )


async def _get_outlook_calendar(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user, required_scope="Calendars.ReadWrite"):
        return g_err
    return await ms.get_outlook_calendar(
        supabase, user, calendar_id=args.get("calendar_id") or ""
    )


async def _list_outlook_events(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user, required_scope="Calendars.ReadWrite"):
        return g_err
    days = max(1, min(int(args.get("days_ahead") or 7), 60))
    now = datetime.now(tz=timezone.utc)
    return {
        "events": await ms.list_outlook_events(
            supabase,
            user,
            calendar_id=args.get("calendar_id"),
            time_min=now - timedelta(hours=2),
            time_max=now + timedelta(days=days),
            limit=int(args.get("limit") or 25),
        )
    }


async def _get_outlook_event(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user, required_scope="Calendars.ReadWrite"):
        return g_err
    return await ms.get_outlook_event(
        supabase, user, event_id=args.get("event_id") or ""
    )


async def _create_outlook_event(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user, required_scope="Calendars.ReadWrite"):
        return g_err
    return await ms.create_outlook_event(
        supabase,
        user,
        subject=args.get("subject") or "",
        start=args.get("start") or "",
        end=args.get("end") or "",
        body=args.get("body"),
        location=args.get("location"),
        attendees=list(args.get("attendees") or []) or None,
        is_all_day=bool(args.get("is_all_day")),
        calendar_id=args.get("calendar_id"),
        online_meeting=bool(args.get("online_meeting")),
    )


async def _update_outlook_event(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user, required_scope="Calendars.ReadWrite"):
        return g_err
    return await ms.update_outlook_event(
        supabase,
        user,
        event_id=args.get("event_id") or "",
        subject=args.get("subject"),
        start=args.get("start"),
        end=args.get("end"),
        body=args.get("body"),
        location=args.get("location"),
        attendees=list(args["attendees"]) if "attendees" in args else None,
    )


async def _delete_outlook_event(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user, required_scope="Calendars.ReadWrite"):
        return g_err
    return await ms.delete_outlook_event(
        supabase, user, event_id=args.get("event_id") or ""
    )


# ---- Outlook contacts ----------------------------------------------------


async def _list_outlook_contacts(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user, required_scope="Contacts.ReadWrite"):
        return g_err
    return {
        "contacts": await ms.list_outlook_contacts(
            supabase, user, limit=int(args.get("limit") or 50)
        )
    }


async def _get_outlook_contact(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user, required_scope="Contacts.ReadWrite"):
        return g_err
    return await ms.get_outlook_contact(
        supabase, user, contact_id=args.get("contact_id") or ""
    )


async def _create_outlook_contact(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user, required_scope="Contacts.ReadWrite"):
        return g_err
    return await ms.create_outlook_contact(
        supabase,
        user,
        name=args.get("name") or "",
        email=args.get("email"),
        phone=args.get("phone"),
        company=args.get("company"),
        job_title=args.get("job_title"),
    )


async def _delete_outlook_contact(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user, required_scope="Contacts.ReadWrite"):
        return g_err
    return await ms.delete_outlook_contact(
        supabase, user, contact_id=args.get("contact_id") or ""
    )


# ---- Outlook attachments -------------------------------------------------


async def _list_outlook_attachments(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    atts = await ms.list_outlook_attachments(
        supabase, user, message_id=args.get("message_id") or ""
    )
    return {"count": len(atts), "attachments": atts}


async def _get_outlook_attachment(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.get_outlook_attachment(
        supabase,
        user,
        message_id=args.get("message_id") or "",
        attachment_id=args.get("attachment_id") or "",
    )


# ---- ToDo list create ----------------------------------------------------


async def _create_todo_list(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    return await ms.create_todo_list(
        supabase, user, display_name=args.get("display_name") or ""
    )


# ---- Meeting prep ----------------------------------------------------


async def _prepare_meeting_brief(args: dict, user: dict, supabase) -> dict:
    event_id = (args.get("event_id") or "").strip()
    source = (args.get("source") or "").strip().lower()
    if not event_id:
        return {"error": "event_id required"}
    if source not in ("google", "outlook"):
        return {"error": "source must be 'google' or 'outlook'"}

    if source == "google":
        if g_err := _need_google(user):
            return g_err
        event = await g.get_calendar_event(supabase, user, event_id=event_id)
        attendee_emails = event.get("attendees") or []
    else:
        if g_err := _need_microsoft(user):
            return g_err
        event = await ms.get_outlook_event(supabase, user, event_id=event_id)
        attendee_emails = event.get("attendees") or []

    # Cap fan-out — a huge meeting invite shouldn't trigger dozens of API calls.
    attendee_emails = attendee_emails[:8]

    attendees: list[dict[str, Any]] = []
    for email in attendee_emails:
        contact = None
        recent_emails: list[dict[str, Any]] = []
        try:
            if source == "google":
                contacts = await g.list_google_contacts(supabase, user, query=email, limit=1)
                contact = contacts[0] if contacts else None
                msgs = await g.list_gmail_messages(
                    supabase, user, limit=5, query=f"from:{email} OR to:{email}"
                )
                recent_emails = [
                    {"subject": m.get("subject"), "snippet": m.get("snippet"), "date": m.get("received_at")}
                    for m in msgs
                ]
            else:
                # No server-side contact search in the Graph wrapper — fetch
                # a batch and match by email client-side.
                all_contacts = await ms.list_outlook_contacts(supabase, user, limit=100)
                contact = next(
                    (c for c in all_contacts if email.lower() in [e.lower() for e in c.get("emails", [])]),
                    None,
                )
                msgs = await ms.list_outlook_messages(
                    supabase, user, limit=5, query=email
                )
                recent_emails = [
                    {"subject": m.get("subject"), "snippet": m.get("snippet"), "date": m.get("received_at")}
                    for m in msgs
                ]
        except Exception:  # noqa: BLE001
            logger.exception("meeting brief lookup failed for %s", email)
        attendees.append({"email": email, "contact": contact, "recent_emails": recent_emails})

    return {"event": event, "attendees": attendees}


# ---- Event-driven triggers -------------------------------------------------


async def _create_email_trigger(args: dict, user: dict, supabase) -> dict:
    name = (args.get("name") or "").strip()
    source = (args.get("source") or "").strip().lower()
    sender_filter = (args.get("sender_filter") or "").strip() or None
    keyword_filter = (args.get("keyword_filter") or "").strip() or None
    prompt = (args.get("prompt") or "").strip()
    channel = (args.get("channel") or "").strip().lower()

    if not name:
        return {"error": "name required"}
    if source not in ("gmail", "outlook"):
        return {"error": "source must be 'gmail' or 'outlook'"}
    if not sender_filter and not keyword_filter:
        return {"error": "at least one of sender_filter or keyword_filter is required"}
    if not prompt:
        return {"error": "prompt required"}
    if channel not in ("web", "email"):
        return {"error": "channel must be 'web', or 'email'"}
    if source == "gmail" and (g_err := _need_google(user)):
        return g_err
    if source == "outlook" and (ms_err := _need_microsoft(user)):
        return ms_err

    try:
        res = (
            supabase.table("email_trigger_flows")
            .insert(
                {
                    "user_id": user["id"],
                    "name": name,
                    "source": source,
                    "sender_filter": sender_filter,
                    "keyword_filter": keyword_filter,
                    "prompt": prompt,
                    "channel": channel,
                }
            )
            .execute()
        )
    except Exception as exc:  # noqa: BLE001
        return {"error": f"could not create trigger: {exc}"}
    row = res.data[0] if res.data else {}
    return {"status": "created", "trigger_id": row.get("id"), "name": name}


async def _list_email_triggers(args: dict, user: dict, supabase) -> dict:
    res = (
        supabase.table("email_trigger_flows")
        .select("id, name, source, sender_filter, keyword_filter, channel, is_active, created_at")
        .eq("user_id", user["id"])
        .order("created_at", desc=True)
        .execute()
    )
    return {"triggers": res.data or []}


async def _delete_email_trigger(args: dict, user: dict, supabase) -> dict:
    trigger_id = (args.get("trigger_id") or "").strip()
    if not trigger_id:
        return {"error": "trigger_id required"}
    res = (
        supabase.table("email_trigger_flows")
        .delete()
        .eq("id", trigger_id)
        .eq("user_id", user["id"])
        .execute()
    )
    if not res.data:
        return {"error": "No trigger found with that id for this user."}
    return {"status": "deleted", "trigger_id": trigger_id}


# ---- Custom commands -----------------------------------------------------

_COMMAND_NAME_RE = re.compile(r"^[a-zA-Z0-9_]{1,32}$")


async def _create_custom_command(args: dict, user: dict, supabase) -> dict:
    name = (args.get("name") or "").strip().lstrip("/").lower()
    prompt_template = (args.get("prompt_template") or "").strip()
    if not name or not _COMMAND_NAME_RE.match(name):
        return {"error": "name must be 1-32 letters/numbers/underscores, no spaces or slash."}
    if not prompt_template:
        return {"error": "prompt_template required"}
    try:
        supabase.table("custom_commands").upsert(
            {"user_id": user["id"], "name": name, "prompt_template": prompt_template},
            on_conflict="user_id,name",
        ).execute()
    except Exception as exc:  # noqa: BLE001
        return {"error": f"could not save command: {exc}"}
    return {"status": "saved", "name": name}


async def _list_custom_commands(args: dict, user: dict, supabase) -> dict:
    res = (
        supabase.table("custom_commands")
        .select("name, prompt_template, created_at")
        .eq("user_id", user["id"])
        .order("name")
        .execute()
    )
    return {"commands": res.data or []}


async def _delete_custom_command(args: dict, user: dict, supabase) -> dict:
    name = (args.get("name") or "").strip().lstrip("/").lower()
    if not name:
        return {"error": "name required"}
    res = (
        supabase.table("custom_commands")
        .delete()
        .eq("user_id", user["id"])
        .eq("name", name)
        .execute()
    )
    if not res.data:
        return {"error": f"No saved command named '{name}'."}
    return {"status": "deleted", "name": name}


# ---- Expense logging ---------------------------------------------------


async def _find_receipt_emails(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    days = min(max(int(args.get("days") or 30), 1), 90)
    limit = min(max(int(args.get("limit") or 20), 1), 30)
    results = await g.find_receipt_emails(supabase, user, days=days, limit=limit)
    return {"count": len(results), "receipts": results}


# ---- Inbox declutter -------------------------------------------------


async def _list_promotional_senders(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    days = min(max(int(args.get("days") or 30), 1), 90)
    return await g.list_promotional_senders(supabase, user, days=days)


async def _declutter_gmail_sender(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    sender_query = (args.get("sender_query") or "").strip()
    action = (args.get("action") or "").strip().lower()
    if not sender_query:
        return {"error": "sender_query required"}
    if action not in ("archive", "trash"):
        return {"error": "action must be 'archive' or 'trash'"}
    days = min(max(int(args.get("days") or 30), 1), 90)
    return await g.declutter_gmail_sender(
        supabase, user, sender_query=sender_query, action=action, days=days
    )


# ---- Memory --------------------------------------------------------------


def _remember_this(args: dict, user: dict, supabase, genai_client) -> dict:
    if not user.get("memory_enabled", True):
        return {"error": "Long-term memory is disabled for this user — nothing was saved."}
    content = (args.get("content") or "").strip()
    if not content:
        return {"error": "content required"}
    kind = (args.get("kind") or "fact").strip().lower()
    if kind not in mem.ALLOWED_KINDS:
        kind = "fact"
    try:
        vec = mem.embed_document(genai_client, content)
    except Exception as exc:  # noqa: BLE001
        logger.exception("remember_this embed failed")
        return {"error": f"Could not save that memory: {exc}"}
    mem.store(supabase, user_id=user["id"], content=content, embedding=vec, kind=kind)
    return {"status": "saved", "content": content, "kind": kind}


def _recall_memory(args: dict, user: dict, supabase, genai_client) -> dict:
    if not user.get("memory_enabled", True):
        return {"error": "Long-term memory is disabled for this user."}
    query = (args.get("query") or "").strip()
    if not query:
        return {"error": "query required"}
    count = min(int(args.get("limit") or 5), 10)
    results = mem.retrieve(
        supabase, genai_client, user_id=user["id"], query=query, count=count
    )
    return {"count": len(results), "memories": results}


def _forget_memory(args: dict, user: dict, supabase) -> dict:
    memory_id = (args.get("memory_id") or "").strip()
    if not memory_id:
        return {"error": "memory_id required"}
    ok = mem.forget(supabase, user_id=user["id"], memory_id=memory_id)
    if not ok:
        return {"error": "No memory found with that id — it may already be gone, or the id wasn't real."}
    return {"status": "forgotten", "memory_id": memory_id}


# ---- Documents (Drive RAG + raw Doc/Sheet/Slides reads) ------------------


def _search_documents(args: dict, user: dict, supabase, genai_client) -> dict:
    query = (args.get("query") or "").strip()
    if not query:
        return {"error": "query required"}
    count = min(int(args.get("limit") or 8), 15)
    
    # Strip outer quotes that the LLM may wrap the query phrase in
    clean_query = query.strip('"').strip("'").strip()
    
    results = doc_rag.search(
        supabase, genai_client, user_id=user["id"], query=clean_query, count=count
    )
    
    # Keyword search fallback if vector search returned nothing
    if not results and len(clean_query) >= 3:
        try:
            import re
            # Parse terms split by 'OR' / 'or'
            terms = [t.strip().strip('"').strip("'").strip() for t in re.split(r'\s+OR\s+|\s+or\s+', clean_query)]
            or_filters = [f"content.ilike.%{t}%" for t in terms if len(t) >= 3]
            if or_filters:
                or_query = ",".join(or_filters)
                res = supabase.table("document_chunks") \
                    .select("chunk_index, content, drive_documents(file_name, drive_file_id)") \
                    .eq("user_id", user["id"]) \
                    .or_(or_query) \
                    .limit(count) \
                    .execute()
                if res.data:
                    results = [
                        {
                            "file_name": r.get("drive_documents", {}).get("file_name") if r.get("drive_documents") else None,
                            "drive_file_id": r.get("drive_documents", {}).get("drive_file_id") if r.get("drive_documents") else None,
                            "chunk_index": r.get("chunk_index"),
                            "similarity": 1.0,
                            "content": r.get("content")
                        }
                        for r in res.data
                    ]
        except Exception as e:
            logger.warning(f"Keyword search fallback failed: {e}")

    if not results:
        return {
            "count": 0,
            "note": "No indexed documents matched. The user may not have indexed any Drive files yet — suggest /dashboard/documents.",
        }
    return {
        "count": len(results),
        "chunks": [
            {
                "file_name": r.get("file_name"),
                "drive_file_id": r.get("drive_file_id"),
                "chunk_index": r.get("chunk_index"),
                "similarity": r.get("similarity"),
                "content": r.get("content"),
            }
            for r in results
        ],
    }


_READ_FULL_DOCUMENT_MAX_CHARS = 24000


def _read_full_document(args: dict, user: dict, supabase) -> dict:
    name = (args.get("file_name") or "").strip()
    if not name:
        return {"error": "file_name required"}
    try:
        docs = (
            supabase.table("drive_documents")
            .select("id, file_name, indexed_at")
            .eq("user_id", user["id"])
            .ilike("file_name", f"%{name}%")
            .order("indexed_at", desc=True)
            .execute()
        )
    except Exception as e:  # noqa: BLE001
        logger.exception("read_full_document lookup failed")
        return {"error": f"lookup failed: {e}"}

    matches = docs.data or []
    if not matches:
        return {
            "count": 0,
            "note": f"No indexed document matches '{name}'. Suggest /dashboard/documents to index it.",
        }
    doc = matches[0]

    chunks_res = (
        supabase.table("document_chunks")
        .select("chunk_index, content")
        .eq("document_id", doc["id"])
        .order("chunk_index")
        .execute()
    )
    chunks = chunks_res.data or []
    if not chunks:
        return {
            "count": 0,
            "note": f"'{doc['file_name']}' is indexed but has no readable text chunks — it may be an image-only scan with no OCR text extracted.",
        }

    full_text = "\n\n".join(c.get("content") or "" for c in chunks)
    truncated = len(full_text) > _READ_FULL_DOCUMENT_MAX_CHARS
    if truncated:
        full_text = full_text[:_READ_FULL_DOCUMENT_MAX_CHARS]

    return {
        "file_name": doc["file_name"],
        "other_matches": [m["file_name"] for m in matches[1:5]] or None,
        "chunk_count": len(chunks),
        "truncated": truncated,
        "content": full_text,
    }


async def _list_drive_files(args: dict, user: dict, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    res = await g.list_drive_files(
        supabase,
        user,
        query=args.get("query") or None,
        folder_id=args.get("folder_id") or None,
        page_size=int(args.get("limit") or 30),
    )
    return res


def _summarize_index(rows: list[dict]) -> dict:
    """Compact summary of an index_folder result for the model."""
    indexed = sum(1 for r in rows if r.get("status") == "indexed")
    failed = [r for r in rows if r.get("status") != "indexed"]
    return {
        "files_total": len(rows),
        "files_indexed": indexed,
        "files_failed": len(failed),
        "chunks_total": sum(int(r.get("chunk_count") or 0) for r in rows),
        "indexed_files": [
            {
                "file_name": r.get("file_name"),
                "chunk_count": r.get("chunk_count", 0),
                "status": r.get("status"),
            }
            for r in rows
        ][:25],
    }


async def _index_drive_file(args: dict, user: dict, genai_client, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    file_id = args.get("file_id") or ""
    if not file_id:
        return {"error": "file_id required"}
    row = await doc_rag.index_file(
        supabase, genai_client, user=user, file_id=file_id, source="gdrive"
    )
    return {
        "status": row.get("status"),
        "file_name": row.get("file_name"),
        "chunk_count": row.get("chunk_count", 0),
        "error": row.get("error"),
    }


async def _index_drive_folder(args: dict, user: dict, genai_client, supabase) -> dict:
    if g_err := _need_google(user):
        return g_err
    folder_id = args.get("folder_id") or ""
    if not folder_id:
        return {"error": "folder_id required"}
    max_files = max(1, min(int(args.get("max_files") or 25), 50))
    rows = await doc_rag.index_folder(
        supabase,
        genai_client,
        user=user,
        folder_id=folder_id,
        max_files=max_files,
        source="gdrive",
    )
    return _summarize_index(rows)


async def _index_onedrive_file(args: dict, user: dict, genai_client, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    file_id = args.get("file_id") or ""
    if not file_id:
        return {"error": "file_id required"}
    row = await doc_rag.index_file(
        supabase, genai_client, user=user, file_id=file_id, source="onedrive"
    )
    return {
        "status": row.get("status"),
        "file_name": row.get("file_name"),
        "chunk_count": row.get("chunk_count", 0),
        "error": row.get("error"),
    }


async def _index_onedrive_folder(args: dict, user: dict, genai_client, supabase) -> dict:
    if g_err := _need_microsoft(user):
        return g_err
    folder_id = args.get("folder_id") or ""
    if not folder_id:
        return {"error": "folder_id required"}
    max_files = max(1, min(int(args.get("max_files") or 25), 50))
    rows = await doc_rag.index_folder(
        supabase,
        genai_client,
        user=user,
        folder_id=folder_id,
        max_files=max_files,
        source="onedrive",
    )
    return _summarize_index(rows)


async def _download_file_for_user(
    args: dict, user: dict, supabase
) -> tuple[Optional[bytes], Optional[str], Optional[str], Optional[dict]]:
    """Common helper: returns (data, filename, mime, error_dict)."""
    file_id = (args.get("file_id") or "").strip()
    source = (args.get("source") or "gdrive").lower()
    if not file_id:
        return None, None, None, {"error": "file_id required"}

    if source == "onedrive":
        if g_err := _need_microsoft(user):
            return None, None, None, g_err
        meta = await ms.get_onedrive_metadata(supabase, user, item_id=file_id)
        name = meta.get("name") or "file"
        mime = meta.get("mime_type") or "application/octet-stream"
        try:
            data = await ms.download_onedrive_file(supabase, user, item_id=file_id)
        except Exception as exc:  # noqa: BLE001
            return None, None, None, {"error": f"onedrive download failed: {exc}"}
        return data, name, mime, None

    # default: gdrive
    if g_err := _need_google(user):
        return None, None, None, g_err
    meta = await g.get_drive_file_metadata(supabase, user, file_id=file_id)
    name = meta.get("name") or "file"
    mime = meta.get("mime_type") or "application/octet-stream"
    # Google-native types (Doc/Sheet/Slide) need export, not download.
    if mime in g.NATIVE_EXPORT:
        export_mime = g.NATIVE_EXPORT[mime]
        try:
            data = await g.export_drive_file(
                supabase, user, file_id=file_id, export_mime=export_mime
            )
        except Exception as exc:  # noqa: BLE001
            return None, None, None, {"error": f"drive export failed: {exc}"}
        # Add suitable extension if missing
        ext = {
            "text/plain": ".txt",
            "text/csv": ".csv",
            "application/pdf": ".pdf",
        }.get(export_mime, "")
        if ext and not name.lower().endswith(ext):
            name = f"{name}{ext}"
        return data, name, export_mime, None
    try:
        data, _ = await g.download_drive_file(supabase, user, file_id=file_id)
    except Exception as exc:  # noqa: BLE001
        return None, None, None, {"error": f"drive download failed: {exc}"}
    return data, name, mime, None


async def _get_pdf_info(args: dict, user: dict, supabase) -> dict:
    data, name, mime, err = await _download_file_for_user(args, user, supabase)
    if err:
        return err
    assert data is not None
    if "pdf" not in (mime or "").lower() and not (name or "").lower().endswith(".pdf"):
        return {"error": f"{name} is not a PDF (mime={mime})"}
    try:
        from pypdf import PdfReader
        from io import BytesIO
        reader = PdfReader(BytesIO(data))
        page_count = len(reader.pages)
        extracted = "".join((p.extract_text() or "") for p in reader.pages[:3])
        has_text = len(extracted.strip()) >= 50
    except Exception as exc:  # noqa: BLE001
        return {"error": f"pdf parse failed: {exc}"}
    return {
        "file_name": name,
        "page_count": page_count,
        "size_bytes": len(data),
        "has_extractable_text": has_text,
        "note": (
            None
            if has_text
            else "This PDF appears to be scanned/image-only. Indexing still "
            "works — we run Gemini OCR automatically as a fallback. Just "
            "call index_drive_file / index_onedrive_file and it will OCR "
            "the pages (may take ~30-60s for long PDFs)."
        ),
    }


async def _get_spreadsheet_info(args: dict, user: dict, supabase) -> dict:
    file_id = (args.get("file_id") or "").strip()
    file_name_query = (args.get("file_name") or "").strip()

    if not file_id and file_name_query:
        res = (
            supabase.table("drive_documents")
            .select("drive_file_id, source, file_name, status")
            .eq("user_id", user["id"])
            .ilike("file_name", f"%{file_name_query}%")
            .order("indexed_at", desc=True)
            .limit(10)
            .execute()
        )
        rows = [r for r in (res.data or []) if r.get("status") == "indexed"]
        if not rows:
            return {
                "error": (
                    f"No indexed spreadsheet matching '{file_name_query}' found. "
                    "It may not be indexed yet — check with search_documents or "
                    "list_drive_files/list_onedrive_files, then index it first."
                )
            }
        distinct = {(r["drive_file_id"], r["source"]) for r in rows}
        if len(distinct) > 1:
            return {
                "error": "multiple indexed files match that name",
                "matches": [
                    {"file_name": r["file_name"], "source": r["source"], "file_id": r["drive_file_id"]}
                    for r in rows
                ],
                "note": "Ask the user which one (they differ by source/id), then retry with file_id+source.",
            }
        args = {**args, "file_id": rows[0]["drive_file_id"], "source": rows[0]["source"]}

    data, name, mime, err = await _download_file_for_user(args, user, supabase)
    if err:
        return err
    assert data is not None
    is_xlsx = (
        "spreadsheetml" in (mime or "").lower()
        or "ms-excel" in (mime or "").lower()
        or (name or "").lower().endswith((".xlsx", ".xls"))
    )
    if not is_xlsx:
        return {
            "error": (
                f"{name} is not a binary Excel file (mime={mime}). If this is "
                "a native Google Sheet, use list_sheet_tabs instead."
            )
        }
    try:
        from openpyxl import load_workbook
        from io import BytesIO
        wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
        sheets = [
            {"title": sheet.title, "row_count": sheet.max_row, "column_count": sheet.max_column}
            for sheet in wb.worksheets
        ]
    except Exception as exc:  # noqa: BLE001
        return {"error": f"xlsx parse failed: {exc}"}
    return {
        "file_name": name,
        "sheet_count": len(sheets),
        "sheets": sheets,
        "size_bytes": len(data),
    }


# Upload bucket name for web-chat file delivery. Created lazily on first use.
_FILE_DELIVERY_BUCKET = "user-deliveries"
_MAX_SEND_BYTES = 50 * 1024 * 1024  # 50 MB


async def _send_file_to_user(args: dict, user: dict, supabase, ctx: dict) -> dict:
    """Deliver a Drive/OneDrive file back to the user in this chat surface."""
    data, name, mime, err = await _download_file_for_user(args, user, supabase)
    if err:
        return err
    assert data is not None and name is not None
    if len(data) > _MAX_SEND_BYTES:
        return {
            "error": (
                f"{name} is {len(data) // (1024 * 1024)} MB which exceeds the "
                f"{_MAX_SEND_BYTES // (1024 * 1024)} MB send limit."
            ),
        }

    surface = (ctx.get("source") or "web").lower()

    return {"error": "Invalid channel. Must be 'web', or 'email'."}
        
    try:
        from croniter import croniter
        croniter(cron_expression)
    except Exception as e:
        return {"error": f"Invalid cron expression: {e}"}
        
    try:
        import pytz
        pytz.timezone(timezone_str)
    except Exception as e:
        return {"error": f"Invalid timezone: {e}"}

    # Guard: check if the user has confirmed in the chat history
    session_id = (context or {}).get("session_id")
    source = (context or {}).get("source")
    if session_id and source:
        try:
            res = (
                supabase.table("messages")
                .select("role, content")
                .eq("user_id", user["id"])
                .eq("source", source)
                .eq("session_id", session_id)
                .order("created_at", desc=True)
                .limit(10)
                .execute()
            )
            history = res.data or []
            history_rev = list(reversed(history))
            
            has_schedule_request = False
            has_confirmation = False
            for msg in history_rev:
                content_lower = (msg.get("content") or "").strip().lower()
                role = msg.get("role")
                if role == "user":
                    if content_lower.startswith(("/schedule", "/shedule", "schedule", "shedule")):
                        has_schedule_request = True
                        has_confirmation = False
                    elif has_schedule_request:
                        conf_words = ["yes", "confirm", "proceed", "yep", "sure", "ok", "go ahead", "y", "do it", "please", "fine", "correct", "agree", "test", "run", "schedule"]
                        if any(w in content_lower for w in conf_words):
                            has_confirmation = True
            
            if has_schedule_request and not has_confirmation:
                return {
                    "error": (
                        "You must first show the user the test results of their schedule prompt, "
                        "and ask them to confirm (e.g. 'Should I schedule this to run automatically?') "
                        "before calling create_scheduled_task. Once the user replies with 'yes' or 'confirm', "
                        "you can call create_scheduled_task in the next turn."
                    )
                }
        except Exception as e:
            logger.exception("Failed to check schedule confirmation in history")

    try:
        data = {
            "user_id": user["id"],
            "name": name,
            "prompt": prompt,
            "cron_expression": cron_expression,
            "timezone": timezone_str,
            "channel": channel,
            "is_active": True
        }
        res = supabase.table("scheduled_tasks").insert(data).execute()
        if res.data:
            return {"success": True, "task": res.data[0]}
        return {"error": "Failed to create scheduled task"}
    except Exception as e:
        logger.exception("create_scheduled_task failed")
        return {"error": str(e)}


async def _list_scheduled_tasks(args: dict, user: dict, supabase) -> dict:
    try:
        res = supabase.table("scheduled_tasks").select("*").eq("user_id", user["id"]).execute()
        return {"tasks": res.data or []}
    except Exception as e:
        logger.exception("list_scheduled_tasks failed")
        return {"error": str(e)}


async def _delete_scheduled_task(args: dict, user: dict, supabase) -> dict:
    task_id = args.get("task_id")
    try:
        res = supabase.table("scheduled_tasks").delete().eq("user_id", user["id"]).eq("id", task_id).execute()
        if res.data:
            return {"success": True, "message": "Scheduled task deleted successfully"}
        return {"error": "Scheduled task not found or not deleted"}
    except Exception as e:
        logger.exception("delete_scheduled_task failed")
        return {"error": str(e)}


async def _create_lead(args: dict, user: dict, supabase) -> dict:
    bot_id = args.get("bot_id")
    if not bot_id:
        return {"error": "bot_id required"}
    session_id = args.get("session_id")
    company = args.get("company") or args.get("company_name")
    fields = {
        "name": args.get("name"),
        "email": args.get("email"),
        "phone": args.get("phone"),
        "company": company,
        "job_title": args.get("job_title"),
        "country": args.get("country"),
        "city": args.get("city"),
        "region": args.get("region"),
        "lat": args.get("lat"),
        "lon": args.get("lon"),
        "industry": args.get("industry"),
        "budget": args.get("budget"),
    }

    standard_keys = [
        "bot_id", "session_id", "name", "email", "phone", "company", "company_name",
        "job_title", "country", "city", "region", "lat", "lon", "industry", "budget",
    ]
    custom_fields = {k: v for k, v in args.items() if k not in standard_keys and v is not None}

    # Dedupe: one lead per widget session. If a lead already exists for this
    # session, MERGE new (non-null) details into it instead of inserting again.
    existing = None
    if session_id:
        try:
            r = supabase.table("chatty_leads").select("*").eq("bot_id", bot_id).eq(
                "session_id", session_id).order("created_at", desc=True).limit(1).execute()
            if r.data:
                existing = r.data[0]
        except Exception:
            logger.exception("lead dedupe lookup failed")

    try:
        if existing:
            update = {k: v for k, v in fields.items() if v}
            if custom_fields:
                update["custom_fields"] = {**(existing.get("custom_fields") or {}), **custom_fields}
            if update:
                supabase.table("chatty_leads").update(update).eq("id", existing["id"]).execute()
            return {"success": True, "lead_id": existing["id"], "message": "Lead updated"}

        insert_data = {"bot_id": bot_id, "session_id": session_id, **fields, "custom_fields": custom_fields}
        res = supabase.table("chatty_leads").insert(insert_data).execute()
        if res.data:
            lead = res.data[0]
            try:
                supabase.table("chatty_audit_logs").insert({
                    "bot_id": bot_id,
                    "action": "lead_created",
                    "details": f"Captured lead: {fields.get('name')} ({fields.get('email')})",
                    "performed_by": "assistant",
                }).execute()
            except Exception:
                pass
            try:
                bot_res = supabase.table("chatty_bots").select("webhook_url").eq("id", bot_id).execute()
                webhook_url = (bot_res.data or [{}])[0].get("webhook_url")
                if webhook_url:
                    await notify.deliver_webhook(
                        url=webhook_url, event="new_lead", bot_id=bot_id,
                        data={k: v for k, v in {**fields, "custom_fields": custom_fields, "lead_id": lead["id"]}.items() if v},
                    )
            except Exception:
                logger.exception("lead webhook delivery failed")
            try:
                await notify.enqueue_webhook_event(
                    supabase, bot_id=bot_id, event="lead.created", session_id=session_id or "",
                    data={"id": lead["id"], **{k: v for k, v in {**fields, "custom_fields": custom_fields}.items() if v}},
                )
            except Exception:
                logger.exception("lead.created webhook enqueue failed")
            return {"success": True, "lead_id": lead["id"], "message": "Lead registered successfully"}
        return {"error": "Failed to create lead"}
    except Exception as e:
        logger.exception("create_lead failed")
        return {"error": str(e)}







# ---------------------------------------------------------------------------
# Social scheduling (Postiz equivalent) — thin wrappers around the same
# tables/helpers main.py's /api/social/* routes use, so chat and the
# dashboard UI stay in sync automatically.
# ---------------------------------------------------------------------------


def _localize_social_datetime(naive_str: str, tz_name: Optional[str]) -> str:
    """Mirrors google_integrations._with_explicit_offset: resolves a possibly
    -naive local datetime string to an unambiguous ISO string with an
    explicit UTC offset, so 'tomorrow at 9am' in the user's timezone lands on
    the right absolute instant regardless of how it's read back later."""
    s = (naive_str or "").strip()
    if not s:
        return s
    if s.endswith("Z") or re.search(r"[+-]\d{2}:?\d{2}$", s):
        return s
    try:
        import pytz
        tz = pytz.timezone(tz_name or "UTC")
        localized = tz.localize(datetime.fromisoformat(s))
        return localized.isoformat()
    except Exception:
        logger.exception("Failed to localize %r to %r for social post", naive_str, tz_name)
        return s


async def _list_connected_social_accounts(args: dict, user: dict, supabase) -> dict:
    import social_providers as sp

    res = (
        supabase.table("social_accounts")
        .select("id, slug, handle, display_name")
        .eq("user_id", user["id"])
        .execute()
    )
    accounts = res.data or []
    connected_slugs = {a["slug"] for a in accounts}
    return {
        # One entry per connected ACCOUNT (not per platform) — a platform can
        # have more than one, e.g. two X accounts, which is why each entry
        # carries its own account_id and handle for disambiguation.
        "accounts": [
            {
                "account_id": a["id"],
                "slug": a["slug"],
                "handle": a.get("handle") or a.get("display_name"),
            }
            for a in accounts
        ],
        "available_to_connect": sorted(
            slug for slug in sp.PROVIDERS_MAP if slug not in connected_slugs and sp.is_real_provider(slug)
        ),
        "note": "Tell the user to connect new platforms from the Social tab's Connected Accounts / Add Channel — you can't connect one for them from chat.",
    }


async def _create_social_post(args: dict, user: dict, supabase) -> dict:
    slug = (args.get("integration_slug") or "").strip().lower()
    account_id = args.get("account_id")
    content = args.get("content") or ""
    if not slug or not content:
        return {"error": "integration_slug and content are required"}

    if account_id:
        acct = (
            supabase.table("social_accounts")
            .select("id, slug")
            .eq("user_id", user["id"])
            .eq("id", account_id)
            .maybe_single()
            .execute()
        )
        if not acct.data:
            return {"error": f"account_id '{account_id}' isn't a connected account for this user."}
        slug = acct.data["slug"]
    else:
        matches = (
            supabase.table("social_accounts")
            .select("id, slug, handle, display_name")
            .eq("user_id", user["id"])
            .eq("slug", slug)
            .execute()
        ).data or []
        if not matches:
            return {"error": f"'{slug}' isn't connected yet. Tell the user to connect it from the Social tab first."}
        if len(matches) > 1:
            options = [{"account_id": m["id"], "handle": m.get("handle") or m.get("display_name")} for m in matches]
            return {
                "error": f"You have {len(matches)} connected '{slug}' accounts — ask the user which one, then retry with account_id set.",
                "options": options,
            }
        account_id = matches[0]["id"]

    publish_date = _localize_social_datetime(args.get("publish_date") or "", user.get("timezone"))
    if not publish_date:
        return {"error": "publish_date is required"}

    row = {
        "user_id": user["id"],
        "integration_slug": slug,
        "social_account_id": account_id,
        "content": content,
        "publish_date": publish_date,
        "state": "draft" if args.get("draft") else "queue",
        "image_url": args.get("image_url"),
    }
    try:
        res = supabase.table("social_posts").insert(row).execute()
        post = res.data[0] if res.data else row
        return {"status": "scheduled", "post_id": post.get("id"), "publish_date": publish_date, "state": row["state"]}
    except Exception as e:
        logger.exception("Failed to create social post from chat")
        return {"error": str(e)}


async def _list_social_posts(args: dict, user: dict, supabase) -> dict:
    limit = min(int(args.get("limit") or 10), 50)
    query = supabase.table("social_posts").select("id, integration_slug, content, publish_date, state, release_url, error").eq("user_id", user["id"])
    state = args.get("state")
    if state:
        query = query.eq("state", state)
    res = query.order("publish_date", desc=True).limit(limit).execute()
    return {"posts": res.data or []}


async def _cancel_social_post(args: dict, user: dict, supabase) -> dict:
    post_id = args.get("post_id")
    if not post_id:
        return {"error": "post_id is required"}
    existing = (
        supabase.table("social_posts")
        .select("state")
        .eq("id", post_id)
        .eq("user_id", user["id"])
        .maybe_single()
        .execute()
    )
    if not existing.data:
        return {"error": "Post not found"}
    if existing.data.get("state") == "published":
        return {"error": "That post is already published and can't be un-published from here."}
    supabase.table("social_posts").delete().eq("id", post_id).eq("user_id", user["id"]).execute()
    return {"status": "cancelled"}


async def _update_social_post(args: dict, user: dict, supabase) -> dict:
    post_id = args.get("post_id")
    if not post_id:
        return {"error": "post_id is required"}
    existing = (
        supabase.table("social_posts")
        .select("state")
        .eq("id", post_id)
        .eq("user_id", user["id"])
        .maybe_single()
        .execute()
    )
    if not existing.data:
        return {"error": "Post not found"}
    if existing.data.get("state") == "published":
        return {"error": "That post is already published and can't be edited from here."}

    update: dict[str, Any] = {}
    if args.get("content"):
        update["content"] = args["content"]
    if args.get("publish_date"):
        publish_date = _localize_social_datetime(args["publish_date"], user.get("timezone"))
        if not publish_date:
            return {"error": "Could not parse publish_date"}
        update["publish_date"] = publish_date
    if "draft" in args and args["draft"] is not None:
        update["state"] = "draft" if args["draft"] else "queue"

    if not update:
        return {"error": "Nothing to update — provide content, publish_date, and/or draft"}

    res = supabase.table("social_posts").update(update).eq("id", post_id).eq("user_id", user["id"]).execute()
    post = res.data[0] if res.data else {}
    return {
        "status": "updated",
        "post_id": post_id,
        "publish_date": post.get("publish_date", update.get("publish_date")),
        "state": post.get("state"),
    }


async def _get_social_analytics(args: dict, user: dict, supabase) -> dict:
    posts_res = supabase.table("social_posts").select("id").eq("user_id", user["id"]).execute()
    post_ids = [p["id"] for p in (posts_res.data or [])]
    if not post_ids:
        return {"impressions": 0, "likes": 0, "reposts": 0, "clicks": 0}
    res = (
        supabase.table("social_analytics")
        .select("impressions, likes, reposts, clicks")
        .in_("post_id", post_ids)
        .execute()
    )
    data = res.data or []
    return {
        "impressions": sum(d.get("impressions") or 0 for d in data),
        "likes": sum(d.get("likes") or 0 for d in data),
        "reposts": sum(d.get("reposts") or 0 for d in data),
        "clicks": sum(d.get("clicks") or 0 for d in data),
    }


# ---------------------------------------------------------------------------
# Web search / page reading — both via Jina AI (s.jina.ai search, r.jina.ai
# reader). Self-contained here rather than importing main.py's existing
# _web_search/_fetch_url_content (used by the embeddable chatty-widget and
# the knowledge-base crawler respectively) to avoid a circular import —
# main.py already imports this module.
# ---------------------------------------------------------------------------


async def _web_search(args: dict, user: dict, supabase) -> dict:
    query = (args.get("query") or "").strip()
    if not query:
        return {"error": "query is required"}
    key = os.environ.get("JINA_API_KEY", "").strip()
    headers = {"Accept": "text/plain", "X-Respond-With": "no-content"}
    if key:
        headers["Authorization"] = f"Bearer {key}"
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            r = await client.get(f"https://s.jina.ai/{urllib.parse.quote(query)}", headers=headers)
            if r.status_code == 200 and r.text.strip():
                return {"results": r.text.strip()[:6000]}
            if r.status_code in (401, 402, 403):
                logger.warning("web_search auth error %s — check JINA_API_KEY", r.status_code)
    except Exception:
        logger.exception("web_search failed for %r", query)
    return {"error": "Web search is unavailable right now."}


async def _read_webpage(args: dict, user: dict, supabase) -> dict:
    url = (args.get("url") or "").strip()
    if not url:
        return {"error": "url is required"}
    headers = {"User-Agent": "KinReader/1.0"}
    key = os.environ.get("JINA_API_KEY", "").strip()
    if key:
        headers["Authorization"] = f"Bearer {key}"
    try:
        async with httpx.AsyncClient(timeout=30, follow_redirects=True, headers=headers) as client:
            r = await client.get(f"https://r.jina.ai/{url}")
            if r.status_code == 429:
                return {"error": "Rate limited fetching that page — try again shortly."}
            if r.status_code >= 300:
                return {"error": f"Could not fetch that page (status {r.status_code})."}
            return {"content": r.text[:12000]}
    except Exception:
        logger.exception("read_webpage failed for %r", url)
        return {"error": "Could not fetch that page."}


# ---------------------------------------------------------------------------
# Voice Agents — chat-driven auto-creation. Only inserts a `draft` row (no
# phone number, no calls placed); the user finishes setup (voice/LLM picks,
# provisioning a number) on the Voice Agents dashboard page. See
# main.py's /api/voice-agents* endpoints for the manual-form equivalent this
# mirrors, and voice_agents.sql for the schema.
# ---------------------------------------------------------------------------

_VOICE_AGENT_USE_CASES = {"sales", "receptionist", "custom"}
_VOICE_AGENT_TOOLS = {"create_calendar_event", "check_calendar_availability", "create_lead"}


async def _create_voice_agent(args: dict, user: dict, supabase) -> dict:
    name = (args.get("name") or "").strip()
    persona = (args.get("persona") or "").strip()
    if not name or not persona:
        return {"error": "name and persona are required"}

    use_case = args.get("use_case") or "custom"
    if use_case not in _VOICE_AGENT_USE_CASES:
        use_case = "custom"

    tools = [t for t in (args.get("tools") or []) if t in _VOICE_AGENT_TOOLS]

    data = {
        "user_id": user["id"],
        "name": name,
        "use_case": use_case,
        "persona": persona,
        "greeting": args.get("greeting"),
        "tools": tools,
        "status": "draft",
        # Fast/cheap defaults — the user can change any of these in the
        # dashboard before provisioning a number.
        "llm_provider": "openai",
        "llm_model": "gpt-4o-mini",
        "stt_provider": "deepgram",
        "tts_provider": "cartesia",
    }
    try:
        res = supabase.table("voice_agents").insert(data).execute()
        if not res.data:
            return {"error": "Failed to create voice agent"}
        row = res.data[0]
        return {
            "id": row["id"],
            "name": row["name"],
            "status": row["status"],
            "dashboard_url": "/dashboard/voice-agents",
        }
    except Exception as e:
        logger.exception("create_voice_agent failed")
        if "duplicate key" in str(e).lower():
            return {"error": f"You already have a voice agent named '{name}'. Pick a different name."}
        return {"error": f"Failed to create voice agent: {e}"}


async def _list_voice_agents(args: dict, user: dict, supabase) -> dict:
    try:
        res = supabase.table("voice_agents").select("id, name, use_case, status, phone_number, created_at").eq("user_id", user["id"]).order("created_at", desc=True).execute()
        return {"voice_agents": res.data or []}
    except Exception as e:
        logger.exception("list_voice_agents failed")
        return {"error": f"Failed to list voice agents: {e}"}


async def _delete_voice_agent(args: dict, user: dict, supabase) -> dict:
    agent_id = args.get("agent_id")
    if not agent_id:
        return {"error": "agent_id is required"}
    try:
        supabase.table("voice_agents").delete().eq("id", agent_id).eq("user_id", user["id"]).execute()
        return {"status": "deleted", "agent_id": agent_id}
    except Exception as e:
        logger.exception("delete_voice_agent failed")
        return {"error": f"Failed to delete voice agent: {e}"}



# ---------------------------------------------------------------------------
# Dispatch
# ---------------------------------------------------------------------------


async def execute(
    name: str,
    args: dict,
    *,
    user: dict,
    supabase,
    genai_client=None,
    context: Optional[dict] = None,
) -> dict:
    """Dispatch a function call.

    `context` is an optional dict that carries surface-specific helpers, e.g.
    {"source": "web"}.
    Tools like send_file_to_user use this to deliver attachments back.
    """
    ctx = context or {}
    try:
        if name.startswith("integration__"):
            parts = name.split("__", 2)
            if len(parts) == 3:
                _, slug, action_name = parts
                # 1. Fetch credentials
                try:
                    cred_res = supabase.table("user_credentials").select("encrypted_payload, auth_type").eq("user_id", user["id"]).eq("integration_slug", slug).maybe_single().execute()
                    cred_data = cred_res.data if cred_res else None
                except Exception as e:
                    logger.exception("Failed to query credentials for custom tool")
                    return {"error": f"Failed to retrieve credentials for {slug}: {e}"}
                
                # Decrypt credentials
                api_key = None
                headers = {}
                if cred_data and cred_data.get("encrypted_payload"):
                    try:
                        enc_bytes = cred_data["encrypted_payload"]
                        if isinstance(enc_bytes, str):
                            if enc_bytes.startswith("\\x") or enc_bytes.startswith(r"\x"):
                                enc_bytes = bytes.fromhex(enc_bytes[2:])
                            else:
                                try:
                                    import base64
                                    enc_bytes = base64.b64decode(enc_bytes)
                                except Exception:
                                    enc_bytes = bytes.fromhex(enc_bytes)
                        elif isinstance(enc_bytes, list):
                            enc_bytes = bytes(enc_bytes)
                            
                        decrypted_text = llm_providers._fernet().decrypt(enc_bytes).decode()
                        creds = json.loads(decrypted_text)
                        api_key = creds.get("api_key")
                    except Exception as e:
                        logger.exception("Failed to decrypt credentials for custom tool")
                        return {"error": f"Failed to decrypt credentials for {slug}: {e}"}
                
                # 2. Fetch integration manifest
                try:
                    integ_res = supabase.table("integrations").select("manifest").eq("slug", slug).maybe_single().execute()
                    manifest = integ_res.data.get("manifest") if integ_res.data else None
                except Exception as e:
                    logger.exception("Failed to query integration manifest")
                    return {"error": f"Failed to retrieve integration manifest for {slug}: {e}"}
                
                if not manifest:
                    return {"error": f"Integration {slug} manifest not found"}
                
                # Find matching action
                action = None
                for act in manifest.get("actions") or []:
                    if act.get("name") == action_name:
                        action = act
                        break
                        
                if not action:
                    return {"error": f"Action {action_name} not found in integration {slug}"}
                
                # 3. Perform request
                path = action.get("path", "")
                method = action.get("method", "GET").upper()
                base_url = manifest.get("base_url") or "https://api.unified.to"
                
                url = f"{base_url.rstrip('/')}/{path.lstrip('/')}"
                query_params = {}
                body_data = {}
                headers = {}
                
                for key, val in args.items():
                    if f"{{{key}}}" in url:
                        url = url.replace(f"{{{key}}}", str(val))
                    else:
                        is_body = False
                        for inp in action.get("inputs") or []:
                            if inp.get("name") == key and inp.get("in") == "body":
                                is_body = True
                                break
                        if is_body:
                            body_data[key] = val
                        else:
                            query_params[key] = val
                            
                if api_key:
                    auth_config = manifest.get("auth", {})
                    auth_type = auth_config.get("type", "api_key")
                    if auth_type == "bearer":
                        headers["Authorization"] = f"Bearer {api_key}"
                    else:
                        header_name = auth_config.get("header_name", "X-API-Key")
                        headers[header_name] = api_key
                        
                import httpx
                async with httpx.AsyncClient() as client:
                    try:
                        req_args = {
                            "method": method,
                            "url": url,
                            "headers": headers,
                            "params": query_params,
                            "timeout": 30.0
                        }
                        if body_data:
                            req_args["json"] = body_data
                            
                        resp = await client.request(**req_args)
                        
                        try:
                            resp_data = resp.json()
                        except Exception:
                            resp_data = {"text": resp.text}
                            
                        return {
                            "status": resp.status_code,
                            "data": resp_data
                        }
                    except Exception as re:
                        logger.exception("HTTP call for custom integration action failed")
                        return {"error": f"HTTP call failed: {re}"}

        if "__" in name:
            server_name, real_tool_name = name.split("__", 1)
            try:
                res = supabase.table("mcp_servers").select("*").eq("user_id", user["id"]).eq("name", server_name).execute()
                server_data = res.data[0] if res.data else None
            except Exception as e:
                logger.exception("Failed to query MCP server info")
                return {"error": f"Failed to retrieve MCP server {server_name}: {e}"}
            if not server_data:
                return {"error": f"MCP server {server_name} not found"}
            import mcp_client
            try:
                headers = await mcp_client.get_mcp_headers(server_data)
                result = await mcp_client.call_remote_tool(
                    sse_url=server_data["url"],
                    tool_name=real_tool_name,
                    arguments=args,
                    headers=headers
                )
                return result
            except Exception as e:
                logger.exception("Remote MCP call failed")
                return {"error": f"Remote MCP call failed: {e}"}

        # Gmail
        if name == "read_gmail":
            return await _read_gmail(args, user, supabase)
        if name == "get_gmail_message":
            return await _get_gmail_message(args, user, supabase)
        if name == "send_email":
            return await _send_email(args, user, supabase)
        if name == "reply_email":
            return await _reply_email(args, user, supabase)
        if name == "draft_email":
            return await _draft_email(args, user, supabase)
        if name == "modify_email_labels":
            return await _modify_email_labels(args, user, supabase)
        if name == "trash_email":
            return await _trash_email(args, user, supabase)
        if name == "list_gmail_labels":
            return await _list_gmail_labels(args, user, supabase)
        if name == "create_gmail_label":
            return await _create_gmail_label(args, user, supabase)
        if name == "delete_email_permanent":
            return await _delete_email_permanent(args, user, supabase)
        # Gmail drafts
        if name == "list_drafts":
            return await _list_drafts(args, user, supabase)
        if name == "get_draft":
            return await _get_draft(args, user, supabase)
        if name == "delete_draft":
            return await _delete_draft(args, user, supabase)
        # Gmail threads
        if name == "list_email_threads":
            return await _list_email_threads(args, user, supabase)
        if name == "get_email_thread":
            return await _get_email_thread(args, user, supabase)
        if name == "reply_to_thread":
            return await _reply_to_thread(args, user, supabase)
        if name == "trash_email_thread":
            return await _trash_email_thread(args, user, supabase)
        if name == "modify_thread_labels":
            return await _modify_thread_labels(args, user, supabase)
        # Calendar
        if name == "read_calendar":
            return await _read_calendar(args, user, supabase)
        if name == "get_calendar_event":
            return await _get_calendar_event(args, user, supabase)
        if name == "create_calendar_event":
            return await _create_calendar_event(args, user, supabase)
        if name == "update_calendar_event":
            return await _update_calendar_event(args, user, supabase)
        if name == "delete_calendar_event":
            return await _delete_calendar_event(args, user, supabase)
        if name == "check_calendar_availability":
            return await _check_calendar_availability(args, user, supabase)
        if name == "create_lead":
            return await _create_lead(args, user, supabase)
        # Social scheduling
        if name == "list_connected_social_accounts":
            return await _list_connected_social_accounts(args, user, supabase)
        if name == "create_social_post":
            return await _create_social_post(args, user, supabase)
        if name == "list_social_posts":
            return await _list_social_posts(args, user, supabase)
        if name == "cancel_social_post":
            return await _cancel_social_post(args, user, supabase)
        if name == "update_social_post":
            return await _update_social_post(args, user, supabase)
        if name == "get_social_analytics":
            return await _get_social_analytics(args, user, supabase)
        # Web search / page reading
        if name == "web_search":
            return await _web_search(args, user, supabase)
        if name == "read_webpage":
            return await _read_webpage(args, user, supabase)
        # Google Tasks
        if name == "list_google_task_lists":
            return await _list_google_task_lists(args, user, supabase)
        if name == "list_google_tasks":
            return await _list_google_tasks(args, user, supabase)
        if name == "get_google_task":
            return await _get_google_task(args, user, supabase)
        if name == "create_google_task":
            return await _create_google_task(args, user, supabase)
        if name == "update_google_task":
            return await _update_google_task(args, user, supabase)
        if name == "delete_google_task":
            return await _delete_google_task(args, user, supabase)
        # Google Contacts
        if name == "list_google_contacts":
            return await _list_google_contacts(args, user, supabase)
        if name == "get_google_contact":
            return await _get_google_contact(args, user, supabase)
        if name == "create_google_contact":
            return await _create_google_contact(args, user, supabase)
        if name == "update_google_contact":
            return await _update_google_contact(args, user, supabase)
        # Voice Agents
        if name == "create_voice_agent":
            return await _create_voice_agent(args, user, supabase)
        if name == "list_voice_agents":
            return await _list_voice_agents(args, user, supabase)
        if name == "delete_voice_agent":
            return await _delete_voice_agent(args, user, supabase)
        # Kin-local
        if name == "read_tasks":
            return _read_tasks(args, user, supabase)
        if name == "read_contacts":
            return _read_contacts(args, user, supabase)
        # Documents (Drive RAG + raw reads)
        if name == "search_documents":
            return _search_documents(args, user, supabase, genai_client)
        if name == "read_full_document":
            return _read_full_document(args, user, supabase)
        if name == "list_drive_files":
            return await _list_drive_files(args, user, supabase)
        if name == "index_drive_file":
            return await _index_drive_file(args, user, genai_client, supabase)
        if name == "index_drive_folder":
            return await _index_drive_folder(args, user, genai_client, supabase)
        if name == "index_onedrive_file":
            return await _index_onedrive_file(args, user, genai_client, supabase)
        if name == "index_onedrive_folder":
            return await _index_onedrive_folder(args, user, genai_client, supabase)
        if name == "get_pdf_info":
            return await _get_pdf_info(args, user, supabase)
        if name == "get_spreadsheet_info":
            return await _get_spreadsheet_info(args, user, supabase)
        if name == "send_file_to_user":
            return await _send_file_to_user(args, user, supabase, ctx)
        if name == "read_google_doc":
            return await _read_google_doc(args, user, supabase)
        if name == "read_google_sheet":
            return await _read_google_sheet(args, user, supabase)
        if name == "list_sheet_tabs":
            return await _list_sheet_tabs(args, user, supabase)
        if name == "read_google_slides":
            return await _read_google_slides(args, user, supabase)
        # Drive writes
        if name == "save_to_drive":
            return await _save_to_drive(args, user, supabase)
        if name == "create_drive_folder":
            return await _create_drive_folder(args, user, supabase)
        if name == "copy_drive_file":
            return await _copy_drive_file(args, user, supabase)
        if name == "move_drive_file":
            return await _move_drive_file(args, user, supabase)
        if name == "rename_drive_file":
            return await _rename_drive_file(args, user, supabase)
        if name == "delete_drive_item":
            return await _delete_drive_item(args, user, supabase)
        if name == "share_drive_item":
            return await _share_drive_item(args, user, supabase)
        # Docs writes
        if name == "create_google_doc":
            return await _create_google_doc(args, user, supabase)
        if name == "append_to_google_doc":
            return await _append_to_google_doc(args, user, supabase)
        if name == "replace_text_in_doc":
            return await _replace_text_in_doc(args, user, supabase)
        # Slides writes
        if name == "replace_text_in_slides":
            return await _replace_text_in_slides(args, user, supabase)
        # Sheets writes
        if name == "create_spreadsheet":
            return await _create_spreadsheet(args, user, supabase)
        if name == "add_sheet_tab":
            return await _add_sheet_tab(args, user, supabase)
        if name == "append_sheet_row":
            return await _append_sheet_row(args, user, supabase)
        if name == "update_sheet_values":
            return await _update_sheet_values(args, user, supabase)
        if name == "clear_sheet_range":
            return await _clear_sheet_range(args, user, supabase)
        # Microsoft Outlook
        if name == "read_outlook":
            return await _read_outlook(args, user, supabase)
        if name == "get_outlook_message":
            return await _get_outlook_message(args, user, supabase)
        if name == "send_outlook_email":
            return await _send_outlook_email(args, user, supabase)
        if name == "reply_outlook_email":
            return await _reply_outlook_email(args, user, supabase)
        if name == "mark_outlook_read":
            return await _mark_outlook_read(args, user, supabase)
        if name == "delete_outlook_message":
            return await _delete_outlook_message(args, user, supabase)
        if name == "move_outlook_message":
            return await _move_outlook_message(args, user, supabase)
        if name == "list_outlook_folders":
            return await _list_outlook_folders(args, user, supabase)
        # Microsoft OneDrive
        if name == "list_onedrive_files":
            return await _list_onedrive_files(args, user, supabase)
        if name == "get_onedrive_metadata":
            return await _get_onedrive_metadata(args, user, supabase)
        # Microsoft ToDo
        if name == "list_todo_lists":
            return await _list_todo_lists(args, user, supabase)
        if name == "list_todo_tasks":
            return await _list_todo_tasks(args, user, supabase)
        if name == "create_todo_task":
            return await _create_todo_task(args, user, supabase)
        if name == "update_todo_task":
            return await _update_todo_task(args, user, supabase)
        if name == "delete_todo_task":
            return await _delete_todo_task(args, user, supabase)
        # OneDrive writes
        if name == "save_to_onedrive":
            return await _save_to_onedrive(args, user, supabase)
        if name == "create_onedrive_folder":
            return await _create_onedrive_folder(args, user, supabase)
        if name == "rename_onedrive_item":
            return await _rename_onedrive_item(args, user, supabase)
        if name == "move_onedrive_item":
            return await _move_onedrive_item(args, user, supabase)
        if name == "copy_onedrive_item":
            return await _copy_onedrive_item(args, user, supabase)
        if name == "delete_onedrive_item":
            return await _delete_onedrive_item(args, user, supabase)
        if name == "share_onedrive_item":
            return await _share_onedrive_item(args, user, supabase)
        # Outlook drafts + message update + folder + folderMessage
        if name == "create_outlook_draft":
            return await _create_outlook_draft(args, user, supabase)
        if name == "update_outlook_draft":
            return await _update_outlook_draft(args, user, supabase)
        if name == "send_outlook_draft":
            return await _send_outlook_draft(args, user, supabase)
        if name == "update_outlook_message":
            return await _update_outlook_message(args, user, supabase)
        if name == "list_outlook_folder_messages":
            return await _list_outlook_folder_messages(args, user, supabase)
        if name == "create_outlook_folder":
            return await _create_outlook_folder(args, user, supabase)
        if name == "get_outlook_folder":
            return await _get_outlook_folder(args, user, supabase)
        # Outlook calendars + events
        if name == "list_outlook_calendars":
            return await _list_outlook_calendars(args, user, supabase)
        if name == "create_outlook_calendar":
            return await _create_outlook_calendar(args, user, supabase)
        if name == "get_outlook_calendar":
            return await _get_outlook_calendar(args, user, supabase)
        if name == "list_outlook_events":
            return await _list_outlook_events(args, user, supabase)
        if name == "get_outlook_event":
            return await _get_outlook_event(args, user, supabase)
        if name == "create_outlook_event":
            return await _create_outlook_event(args, user, supabase)
        if name == "update_outlook_event":
            return await _update_outlook_event(args, user, supabase)
        if name == "delete_outlook_event":
            return await _delete_outlook_event(args, user, supabase)
        # Outlook contacts
        if name == "list_outlook_contacts":
            return await _list_outlook_contacts(args, user, supabase)
        if name == "get_outlook_contact":
            return await _get_outlook_contact(args, user, supabase)
        if name == "create_outlook_contact":
            return await _create_outlook_contact(args, user, supabase)
        if name == "delete_outlook_contact":
            return await _delete_outlook_contact(args, user, supabase)
        # Outlook attachments
        if name == "list_outlook_attachments":
            return await _list_outlook_attachments(args, user, supabase)
        if name == "get_outlook_attachment":
            return await _get_outlook_attachment(args, user, supabase)
        # ToDo list create
        if name == "create_todo_list":
            return await _create_todo_list(args, user, supabase)
        # Scheduled Tasks
        if name == "create_scheduled_task":
            return await _create_scheduled_task(args, user, supabase, ctx)
        if name == "list_scheduled_tasks":
            return await _list_scheduled_tasks(args, user, supabase)
        if name == "delete_scheduled_task":
            return await _delete_scheduled_task(args, user, supabase)
        # Meeting prep
        if name == "prepare_meeting_brief":
            return await _prepare_meeting_brief(args, user, supabase)
        # Expense logging
        if name == "find_receipt_emails":
            return await _find_receipt_emails(args, user, supabase)
        # Event-driven triggers
        if name == "create_email_trigger":
            return await _create_email_trigger(args, user, supabase)
        if name == "list_email_triggers":
            return await _list_email_triggers(args, user, supabase)
        if name == "delete_email_trigger":
            return await _delete_email_trigger(args, user, supabase)
        # Custom commands
        if name == "create_custom_command":
            return await _create_custom_command(args, user, supabase)
        if name == "list_custom_commands":
            return await _list_custom_commands(args, user, supabase)
        if name == "delete_custom_command":
            return await _delete_custom_command(args, user, supabase)
        # Inbox declutter
        if name == "list_promotional_senders":
            return await _list_promotional_senders(args, user, supabase)
        if name == "declutter_gmail_sender":
            return await _declutter_gmail_sender(args, user, supabase)
        # Email follow-ups
        if name == "send_followup_nudge":
            return await _send_followup_nudge(args, user, supabase)
        # Memory
        if name == "remember_this":
            return _remember_this(args, user, supabase, genai_client)
        if name == "recall_memory":
            return _recall_memory(args, user, supabase, genai_client)
        if name == "forget_memory":
            return _forget_memory(args, user, supabase)
        return {"error": f"unknown tool: {name}"}
    except ms.MicrosoftNotConnected:
        return {
            "error": "Microsoft not connected. Ask the user to visit /dashboard/integrations.",
        }
    except g.GoogleNotConnected:
        return {
            "error": "Google not connected. Ask the user to visit /dashboard/integrations to connect their Google account.",
        }
    except Exception as exc:  # noqa: BLE001
        logger.exception("tool %s failed", name)
        return {"error": f"tool failed: {exc}"}
