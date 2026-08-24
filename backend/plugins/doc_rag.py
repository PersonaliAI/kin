"""Document RAG — index Google Drive files into pgvector and search them.

Pipeline:
  1. extract_text(file) — pulls plain text from a Drive file:
       Google Doc      → export text/plain
       Google Sheet    → export text/csv (or sheets API)
       Google Slides   → export text/plain
       PDF             → download bytes + markitdown (pypdf fallback)
       DOCX            → download bytes + markitdown (python-docx fallback)
       XLSX            → download bytes + markitdown (openpyxl fallback)
       text/markdown   → download as-is
  2. chunk_text(text) — recursive split on paragraphs/sentences/words
       Target ~2000 chars per chunk, 200 char overlap.
  3. embed each chunk (text-embedding-004, 768-d)
  4. UPSERT into drive_documents + document_chunks

Text extraction (stage 3): native PDF/DOCX/XLSX text+structure extraction
goes through Microsoft's ``markitdown`` library first (see
``_markitdown_convert``), which reads whatever text layer is already in the
file. The legacy pypdf/python-docx/openpyxl extractors are kept as a
best-effort fallback if markitdown isn't installed or returns nothing (e.g.
an unusual file that trips up one converter but not the other). markitdown
does NOT do OCR — it can't read text out of scanned pages or photos, so
Gemini vision OCR (``ocr_image`` / ``_ocr_pdf_with_gemini``) remains the
fallback for genuinely visual/scanned content, unchanged in that respect.

Gemini calls (stage 3 of the litellm migration): ``ocr_image`` now goes
through ``app/core/llm.py``'s unified ``complete()`` instead of calling
``google-genai`` directly. ``_ocr_pdf_with_gemini`` (PDF-native OCR, which
sends whole PDF byte blobs rather than a single image) was intentionally
left on the direct ``google-genai`` client for this stage — litellm's
OpenAI-compatible message shape doesn't have a well-established
document/PDF content-part convention the way it does for images, so
migrating it needs its own verification pass rather than reusing the same
pattern as image OCR.
"""

from __future__ import annotations

import base64
import io
import logging
import os
import re
from typing import Any, Optional

from google import genai
from google.genai import types as genai_types

from app.core.llm import complete as llm_complete
from plugins import google_integrations as g
from plugins import memory as mem
from plugins import microsoft_integrations as ms

logger = logging.getLogger("kin.doc_rag")

CHUNK_SIZE = 2000
CHUNK_OVERLAP = 200
MAX_CHARS_PER_DOC = 200_000  # ~50k tokens — hard cap to avoid runaway costs

# OCR: Gemini accepts PDFs natively up to ~3600 pages, but we cap so we don't
# burn the user's tokens on a 1000-page scan. 100 pages * 258 input tokens =
# ~26k tokens — well within free-tier budget.
OCR_MAX_PAGES = int(os.environ.get("KIN_OCR_MAX_PAGES", "100"))
OCR_MODEL = os.environ.get("KIN_OCR_MODEL", "gemini-2.5-flash")


# ---------------------------------------------------------------------------
# Text extraction
# ---------------------------------------------------------------------------


_markitdown_instance: Any = None


def _get_markitdown() -> Any:
    """Lazily construct a single shared MarkItDown instance (it loads a small
    file-type-detection model on init, so don't rebuild it per call)."""
    global _markitdown_instance
    if _markitdown_instance is None:
        from markitdown import MarkItDown

        _markitdown_instance = MarkItDown()
    return _markitdown_instance


def _markitdown_convert(data: bytes, *, extension: str, label: str = "") -> str:
    """Convert a document blob to text via markitdown. Returns '' on any
    failure (not installed, unsupported/corrupt file, no text found) so
    callers can fall back to the legacy format-specific extractor.

    NOTE: markitdown's PDF/image handling reads an existing text layer only
    — it does not perform OCR on scanned pages or photos. That's still
    Gemini vision's job (see ocr_image / _ocr_pdf_with_gemini below).
    """
    try:
        md = _get_markitdown()
        result = md.convert_stream(io.BytesIO(data), file_extension=extension)
        return (result.text_content or "").strip()
    except ImportError:
        logger.warning("markitdown not installed; falling back to legacy extractor for %s", label or extension)
        return ""
    except Exception as exc:  # noqa: BLE001
        logger.info("markitdown conversion failed for %s (%s): %s", label or "blob", extension, exc)
        return ""


def _extract_pdf_text(data: bytes) -> str:
    """Pull text from a PDF's native text layer — markitdown first, pypdf as
    a fallback. Returns '' for scanned/image-only PDFs (no text layer at all).

    Sync. For full OCR fallback, use `extract_pdf_text_with_ocr` instead.
    """
    text = _markitdown_convert(data, extension=".pdf", label="pdf")
    if text:
        return text

    try:
        from pypdf import PdfReader
    except ImportError:
        logger.warning("pypdf not installed; PDF extraction unavailable")
        return ""
    try:
        reader = PdfReader(io.BytesIO(data))
        text = "\n\n".join(p.extract_text() or "" for p in reader.pages)
        if reader.pages and len(text.strip()) < 20:
            logger.info(
                "PDF parsed (%d pages) but contains no extractable text — "
                "likely scanned/image-only.",
                len(reader.pages),
            )
            return ""
        return text
    except Exception as exc:  # noqa: BLE001
        logger.warning("PDF parse failed: %s", exc)
        return ""


def _split_pdf_pages(data: bytes, max_pages: int) -> list[bytes]:
    """Slice a PDF into one or more PDF byte-blobs of at most `max_pages` each.

    Gemini caps PDF input around ~3600 pages but practical limits (token cost,
    response size, timeouts) push us to chunk at ~100 pages. We split by
    creating new mini-PDFs so each Gemini call gets a manageable slice.
    """
    try:
        from pypdf import PdfReader, PdfWriter
    except ImportError:
        return [data]  # best-effort: ship the whole thing
    try:
        reader = PdfReader(io.BytesIO(data))
        total = len(reader.pages)
        if total <= max_pages:
            return [data]
        chunks: list[bytes] = []
        for start in range(0, total, max_pages):
            writer = PdfWriter()
            for idx in range(start, min(start + max_pages, total)):
                writer.add_page(reader.pages[idx])
            buf = io.BytesIO()
            writer.write(buf)
            chunks.append(buf.getvalue())
        logger.info("PDF split into %d chunks of <=%d pages each", len(chunks), max_pages)
        return chunks
    except Exception as exc:  # noqa: BLE001
        logger.warning("PDF page-split failed: %s — sending whole PDF", exc)
        return [data]


async def _ocr_pdf_with_gemini(
    client: genai.Client, data: bytes, *, label: str = "pdf"
) -> str:
    """Run OCR on a scanned PDF via Gemini's multimodal vision.

    Sends PDF bytes directly as a Part — Gemini natively handles PDF input
    and reads text from page images. Cheaper and more accurate than running
    Tesseract on rendered images.
    """
    OCR_PROMPT = (
        "Extract ALL readable text from this PDF, in reading order. "
        "Output plain text only — no commentary, no markdown, no JSON. "
        "Preserve paragraph breaks. If a page has tables, render them as "
        "tab-separated rows. Do NOT skip pages."
    )
    pieces = _split_pdf_pages(data, OCR_MAX_PAGES)
    extracted: list[str] = []
    for i, piece in enumerate(pieces):
        try:
            resp = await client.aio.models.generate_content(
                model=OCR_MODEL,
                contents=[
                    genai_types.Content(
                        role="user",
                        parts=[
                            genai_types.Part.from_bytes(
                                data=piece, mime_type="application/pdf"
                            ),
                            genai_types.Part.from_text(text=OCR_PROMPT),
                        ],
                    )
                ],
                config=genai_types.GenerateContentConfig(
                    temperature=0,
                    max_output_tokens=8192,
                ),
            )
            text = (resp.text or "").strip()
            if text:
                extracted.append(text)
                logger.info(
                    "OCR slice %d/%d of %s: %d chars",
                    i + 1, len(pieces), label, len(text),
                )
            else:
                logger.warning("OCR slice %d/%d of %s returned no text", i + 1, len(pieces), label)
        except Exception as exc:  # noqa: BLE001
            logger.warning("OCR slice %d/%d of %s failed: %s", i + 1, len(pieces), label, exc)
    return "\n\n".join(extracted)


async def ocr_image(
    data: bytes, *, mime_type: str, label: str = "image", user_id: Optional[str] = None
) -> str:
    """Extract readable text from a single standalone photo (not a PDF page)
    via Gemini vision — e.g. a job ad, whiteboard, receipt, or screenshot
    someone sends directly in web chat, as opposed to a scanned PDF.

    Stage 3: migrated onto app/core/llm.py's unified `complete()`. Vision
    input is expressed as an OpenAI-style multimodal message — a list of
    text/image_url content parts, with the image base64-inlined as a data
    URI — which litellm translates to Gemini's native inline-image format.
    """
    OCR_PROMPT = (
        "Extract ALL readable text from this image, in reading order. "
        "Output plain text only — no commentary, no markdown, no JSON. "
        "Preserve paragraph/line breaks. If it's a table, render it as "
        "tab-separated rows. If there is no readable text at all, output "
        "nothing."
    )
    mime = mime_type or "image/jpeg"
    b64 = base64.b64encode(data).decode("ascii")
    messages = [
        {
            "role": "user",
            "content": [
                {"type": "text", "text": OCR_PROMPT},
                {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}},
            ],
        }
    ]
    try:
        result = await llm_complete(
            model=OCR_MODEL,
            messages=messages,
            temperature=0,
            max_tokens=8192,
            feature="doc_ocr",
            user_id=user_id,
        )
        text = (result.text or "").strip()
        logger.info("image OCR of %s: %d chars", label, len(text))
        return text
    except Exception as exc:  # noqa: BLE001
        logger.warning("image OCR failed for %s: %s", label, exc)
        return ""


async def extract_pdf_text_with_ocr(
    client: genai.Client, data: bytes, *, label: str = "pdf"
) -> str:
    """Get text from a PDF — pypdf first, OCR fallback for scanned PDFs.

    Use this instead of `_extract_pdf_text` whenever you have a genai client
    available. Returns the extracted text, capped at MAX_CHARS_PER_DOC.
    """
    fast = _extract_pdf_text(data)
    if fast.strip():
        return fast[:MAX_CHARS_PER_DOC]
    # Scanned PDF — fall back to Gemini OCR.
    logger.info("Falling back to Gemini OCR for %s", label)
    ocred = await _ocr_pdf_with_gemini(client, data, label=label)
    return ocred[:MAX_CHARS_PER_DOC]


def _extract_docx_text(data: bytes) -> str:
    """Pull text from a DOCX — markitdown (via mammoth) first, python-docx
    as a fallback if markitdown isn't installed or fails on this file."""
    text = _markitdown_convert(data, extension=".docx", label="docx")
    if text:
        return text

    try:
        from docx import Document
    except ImportError:
        logger.warning("python-docx not installed; DOCX extraction unavailable")
        return ""
    try:
        doc = Document(io.BytesIO(data))
        return "\n".join(p.text for p in doc.paragraphs if p.text)
    except Exception as exc:  # noqa: BLE001
        logger.warning("DOCX parse failed: %s", exc)
        return ""


def _extract_xlsx_text(data: bytes) -> str:
    """Pull a plain-text table dump from an uploaded .xlsx (not a native
    Google Sheet — those go through NATIVE_EXPORT as CSV instead).

    markitdown first, openpyxl as a fallback if markitdown isn't installed
    or fails on this file.
    """
    text = _markitdown_convert(data, extension=".xlsx", label="xlsx")
    if text:
        return text

    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not installed; XLSX extraction unavailable")
        return ""
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        parts: list[str] = []
        for sheet in wb.worksheets:
            rows_text: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    rows_text.append("\t".join(cells))
            if rows_text:
                parts.append(f"# Sheet: {sheet.title}\n" + "\n".join(rows_text))
        return "\n\n".join(parts)
    except Exception as exc:  # noqa: BLE001
        logger.warning("XLSX parse failed: %s", exc)
        return ""


async def extract_text_onedrive(
    supabase,
    user: dict[str, Any],
    *,
    file_id: str,
    mime_type: str,
    genai_client: Optional[genai.Client] = None,
) -> str:
    """Pull plain text from a OneDrive file. Uses OCR fallback for scanned PDFs."""
    try:
        data = await ms.download_onedrive_file(supabase, user, item_id=file_id)
    except Exception as exc:  # noqa: BLE001
        logger.warning("OneDrive download failed for %s: %s", file_id, exc)
        return ""
    if mime_type == "application/pdf":
        if genai_client is not None:
            return await extract_pdf_text_with_ocr(genai_client, data, label=file_id)
        return _extract_pdf_text(data)[:MAX_CHARS_PER_DOC]
    if mime_type in (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    ):
        return _extract_docx_text(data)[:MAX_CHARS_PER_DOC]
    if mime_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        return _extract_xlsx_text(data)[:MAX_CHARS_PER_DOC]
    if mime_type and mime_type.startswith("text/"):
        try:
            return data.decode("utf-8", errors="replace")[:MAX_CHARS_PER_DOC]
        except Exception:  # noqa: BLE001
            return ""
    return ""


async def extract_text(
    supabase,
    user: dict[str, Any],
    *,
    file_id: str,
    mime_type: str,
    genai_client: Optional[genai.Client] = None,
) -> str:
    """Pull plain text from any supported Drive file. Returns empty string if unsupported."""
    # Google-native types — export via Drive API.
    if mime_type in g.NATIVE_EXPORT:
        try:
            data = await g.export_drive_file(
                supabase, user, file_id=file_id, export_mime=g.NATIVE_EXPORT[mime_type]
            )
            return data.decode("utf-8", errors="replace")[:MAX_CHARS_PER_DOC]
        except Exception as exc:  # noqa: BLE001
            logger.warning("native export failed for %s: %s", file_id, exc)
            return ""

    # Binary downloads
    if mime_type in g.SUPPORTED_BINARY_MIMES or mime_type.startswith("text/"):
        try:
            data, _ = await g.download_drive_file(supabase, user, file_id=file_id)
        except Exception as exc:  # noqa: BLE001
            logger.warning("download failed for %s: %s", file_id, exc)
            return ""

        if mime_type == "application/pdf":
            if genai_client is not None:
                return await extract_pdf_text_with_ocr(
                    genai_client, data, label=file_id
                )
            return _extract_pdf_text(data)[:MAX_CHARS_PER_DOC]
        if mime_type in (
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/msword",
        ):
            return _extract_docx_text(data)[:MAX_CHARS_PER_DOC]
        if mime_type in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ):
            return _extract_xlsx_text(data)[:MAX_CHARS_PER_DOC]
        # plain text / csv / md
        try:
            return data.decode("utf-8", errors="replace")[:MAX_CHARS_PER_DOC]
        except Exception:  # noqa: BLE001
            return ""

    # Unsupported (images, video, etc.)
    return ""


# ---------------------------------------------------------------------------
# Chunking
# ---------------------------------------------------------------------------


_SPLIT_PATTERNS = [
    r"\n\n+",  # paragraphs
    r"\n",
    r"(?<=[.!?])\s+",  # sentences
    r"\s+",  # words
]


def chunk_text(
    text: str, size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP
) -> list[str]:
    """Recursive character splitter — tries to break on paragraph/sentence boundaries."""
    text = (text or "").strip()
    if not text:
        return []
    if len(text) <= size:
        return [text]

    chunks: list[str] = []
    start = 0
    while start < len(text):
        end = min(start + size, len(text))
        if end < len(text):
            # Try to back off to a natural boundary in the last 30% of the window.
            window = text[start + int(size * 0.7) : end]
            best_break = -1
            for pat in _SPLIT_PATTERNS:
                m = list(re.finditer(pat, window))
                if m:
                    best_break = start + int(size * 0.7) + m[-1].end()
                    break
            if best_break > start:
                end = best_break
        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)
        if end >= len(text):
            break
        start = max(end - overlap, start + 1)
    return chunks


# ---------------------------------------------------------------------------
# Index a single file
# ---------------------------------------------------------------------------


async def index_file(
    supabase,
    genai_client: genai.Client,
    *,
    user: dict[str, Any],
    file_id: str,
    file_meta: Optional[dict[str, Any]] = None,
    source: str = "gdrive",
) -> dict[str, Any]:
    """Fetch + extract + chunk + embed + store one file. Returns the doc row.

    source: "gdrive" or "onedrive".
    """
    user_id = user["id"]
    if source == "onedrive":
        meta = file_meta or await ms.get_onedrive_metadata(
            supabase, user, item_id=file_id
        )
    else:
        meta = file_meta or await g.get_drive_file_metadata(
            supabase, user, file_id=file_id
        )
    name = meta.get("name") or "(untitled)"
    mime = meta.get("mime_type") or ""

    # Insert/upsert the document row in 'pending' state first.
    upsert_payload = {
        "user_id": user_id,
        "drive_file_id": file_id,
        "file_name": name,
        "mime_type": mime,
        "web_view_link": meta.get("web_view_link"),
        "parent_folder_id": meta.get("parent_folder_id"),
        "size_bytes": meta.get("size_bytes"),
        "modified_at": meta.get("modified_at"),
        "source": source,
        "status": "pending",
        "error": None,
    }
    row_res = (
        supabase.table("drive_documents")
        .upsert(upsert_payload, on_conflict="user_id,drive_file_id")
        .execute()
    )
    doc_row = row_res.data[0] if row_res.data else None
    if not doc_row:
        # Race — fetch it
        existing = (
            supabase.table("drive_documents")
            .select("*")
            .eq("user_id", user_id)
            .eq("drive_file_id", file_id)
            .single()
            .execute()
        )
        doc_row = existing.data

    document_id = doc_row["id"]

    # Wipe any prior chunks for this document — we reindex from scratch.
    supabase.table("document_chunks").delete().eq(
        "document_id", document_id
    ).execute()

    try:
        if source == "onedrive":
            text = await extract_text_onedrive(
                supabase, user, file_id=file_id, mime_type=mime,
                genai_client=genai_client,
            )
        else:
            text = await extract_text(
                supabase, user, file_id=file_id, mime_type=mime,
                genai_client=genai_client,
            )
        if not text.strip():
            err = (
                "OCR ran but recovered no readable text. The pages may be "
                "blank, photo-only with no visible text, or in a script the "
                "model couldn't read. Try uploading a different copy."
                if (mime or "").lower() == "application/pdf"
                else "no extractable text from this file type."
            )
            supabase.table("drive_documents").update(
                {
                    "status": "failed",
                    "error": err,
                    "chunk_count": 0,
                }
            ).eq("id", document_id).execute()
            return {**doc_row, "status": "failed", "error": err}

        chunks = chunk_text(text)
        rows: list[dict[str, Any]] = []
        if chunks:
            # Filename becomes the "title" prefix for v2 — improves retrieval.
            titles = [name] * len(chunks)
            try:
                vectors = mem.embed_documents_batch(
                    genai_client, chunks, titles=titles, batch_size=50
                )
            except Exception as exc:  # noqa: BLE001
                logger.warning(
                    "batch embed failed for %s: %s — falling back to per-chunk", name, exc
                )
                vectors = []
                for idx, c in enumerate(chunks):
                    try:
                        vectors.append(
                            mem.embed_document(genai_client, c, title=name)
                        )
                    except Exception as e2:  # noqa: BLE001
                        logger.warning("embed failed (chunk %d of %s): %s", idx, name, e2)
                        vectors.append(None)  # placeholder so indexes align

            for idx, (c, vec) in enumerate(zip(chunks, vectors)):
                if vec is None:
                    continue
                rows.append(
                    {
                        "user_id": user_id,
                        "document_id": document_id,
                        "drive_file_id": file_id,
                        "file_name": name,
                        "chunk_index": idx,
                        "content": c,
                        "embedding": vec,
                        "metadata": {"mime_type": mime},
                    }
                )
        if rows:
            # Batch insert (Postgrest can handle a few hundred rows per call)
            for batch_start in range(0, len(rows), 50):
                supabase.table("document_chunks").insert(
                    rows[batch_start : batch_start + 50]
                ).execute()
            final_status = "indexed"
            err_msg = None
        else:
            # We extracted text but no chunks survived embedding — mark failed
            # so the UI shows the problem instead of silently saying 0 chunks.
            final_status = "failed"
            err_msg = "embeddings failed for all chunks (check model availability)"

        supabase.table("drive_documents").update(
            {
                "status": final_status,
                "error": err_msg,
                "chunk_count": len(rows),
                "indexed_at": "now()",
            }
        ).eq("id", document_id).execute()
        return {**doc_row, "status": final_status, "chunk_count": len(rows)}

    except Exception as exc:  # noqa: BLE001
        logger.exception("indexing failed for %s", file_id)
        supabase.table("drive_documents").update(
            {"status": "failed", "error": str(exc)[:500]}
        ).eq("id", document_id).execute()
        return {**doc_row, "status": "failed", "error": str(exc)}


async def _extract_text_from_blob(
    data: bytes,
    mime_type: str,
    filename: str,
    *,
    genai_client: Optional[genai.Client] = None,
    user_id: Optional[str] = None,
) -> str:
    """Pull plain text from raw bytes — handles uploads (no Drive backing).

    PDFs get OCR fallback when `genai_client` is provided. `genai_client`
    being non-None is used purely as the "OCR capability available" signal
    here (kept for parity with the rest of this module's call sites) — the
    actual OCR call goes through app/core/llm.py's `complete()`, not the
    genai client directly. `user_id` (if known) is threaded through to the
    llm_calls cost/usage ledger.
    """
    m = (mime_type or "").lower()
    name_lower = (filename or "").lower()

    if m == "application/pdf" or name_lower.endswith(".pdf"):
        if genai_client is not None:
            return await extract_pdf_text_with_ocr(
                genai_client, data, label=filename or "upload.pdf"
            )
        return _extract_pdf_text(data)[:MAX_CHARS_PER_DOC]
    if (
        m in (
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/msword",
        )
        or name_lower.endswith((".docx", ".doc"))
    ):
        return _extract_docx_text(data)[:MAX_CHARS_PER_DOC]
    # Plain text family: txt/md/csv/json/etc.
    if m.startswith("text/") or name_lower.endswith(
        (".txt", ".md", ".markdown", ".csv", ".tsv", ".json", ".log")
    ):
        try:
            return data.decode("utf-8", errors="replace")[:MAX_CHARS_PER_DOC]
        except Exception:  # noqa: BLE001
            return ""
    if m.startswith("image/") or name_lower.endswith(
        (".jpg", ".jpeg", ".png", ".webp", ".heic", ".heif")
    ):
        if genai_client is not None:
            return (
                await ocr_image(data, mime_type=m or "image/jpeg", label=filename, user_id=user_id)
            )[:MAX_CHARS_PER_DOC]
        return ""
    return ""


_DOCS_BUCKET = "kin-documents"


def _store_original_bytes(supabase, *, user_id: str, document_id: str, data: bytes) -> Optional[str]:
    """Best-effort upload of the raw file bytes so it can be attached to an
    email later. Never raises — a storage hiccup shouldn't fail indexing,
    it just means this particular file won't be attachable afterward."""
    try:
        try:
            supabase.storage.create_bucket(_DOCS_BUCKET, options={"public": False})
        except Exception:  # noqa: BLE001
            pass  # already exists
        path = f"{user_id}/{document_id}"
        supabase.storage.from_(_DOCS_BUCKET).upload(
            path, data, {"content-type": "application/octet-stream", "upsert": "true"}
        )
        return path
    except Exception:  # noqa: BLE001
        logger.exception("failed to store original bytes for document %s", document_id)
        return None


async def get_document_bytes(supabase, user: dict[str, Any], file_name: str) -> dict[str, Any]:
    """Fetch a previously-indexed document's ORIGINAL file bytes (not the
    extracted text) by filename, for attaching to an outgoing email.
    Drive/OneDrive-sourced files are re-downloaded live via drive_file_id;
    direct uploads come from the kin-documents storage bucket."""
    res = (
        supabase.table("drive_documents")
        .select("id, file_name, mime_type, source, drive_file_id, storage_path")
        .eq("user_id", user["id"])
        .ilike("file_name", f"%{file_name}%")
        .order("indexed_at", desc=True)
        .execute()
    )
    matches = res.data or []
    if not matches:
        return {"error": f"No indexed document matches '{file_name}'."}
    doc = matches[0]

    if doc.get("storage_path"):
        try:
            data = supabase.storage.from_(_DOCS_BUCKET).download(doc["storage_path"])
        except Exception as exc:  # noqa: BLE001
            return {"error": f"Could not retrieve the stored file: {exc}"}
        return {"file_name": doc["file_name"], "mime_type": doc.get("mime_type"), "data": data}

    if doc.get("source") == "gdrive" and doc.get("drive_file_id"):
        data, mime = await g.download_drive_file(supabase, user, file_id=doc["drive_file_id"])
        return {"file_name": doc["file_name"], "mime_type": mime or doc.get("mime_type"), "data": data}

    if doc.get("source") == "onedrive" and doc.get("drive_file_id"):
        data = await ms.download_onedrive_file(supabase, user, item_id=doc["drive_file_id"])
        return {"file_name": doc["file_name"], "mime_type": doc.get("mime_type"), "data": data}

    return {
        "error": f"'{doc['file_name']}' was indexed before attachment support was added — "
        "re-upload or re-index it to attach it to an email."
    }


async def index_blob(
    supabase,
    genai_client: genai.Client,
    *,
    user: dict[str, Any],
    filename: str,
    data: bytes,
    mime_type: str,
    source: str = "upload",
    external_ref: Optional[str] = None,
) -> dict[str, Any]:
    """Index a file from raw bytes (no Drive/OneDrive needed).

    Used by:
      - POST /api/documents/upload (web chat paperclip)

    `external_ref` is a stable id used for de-dup. If omitted we hash the
    filename + size so a re-upload of the same file replaces existing chunks
    instead of duplicating.
    """
    import hashlib
    user_id = user["id"]

    name = (filename or "untitled").strip() or "untitled"
    ref = external_ref or (
        "upload-" + hashlib.sha1(f"{name}|{len(data)}".encode()).hexdigest()[:24]
    )

    upsert_payload = {
        "user_id": user_id,
        "drive_file_id": ref,  # reuse the unique slot
        "file_name": name,
        "mime_type": mime_type or "application/octet-stream",
        "web_view_link": None,
        "parent_folder_id": None,
        "size_bytes": len(data),
        "modified_at": None,
        "source": source,
        "status": "pending",
        "error": None,
    }
    row_res = (
        supabase.table("drive_documents")
        .upsert(upsert_payload, on_conflict="user_id,drive_file_id")
        .execute()
    )
    doc_row = row_res.data[0] if row_res.data else None
    if not doc_row:
        existing = (
            supabase.table("drive_documents")
            .select("*")
            .eq("user_id", user_id)
            .eq("drive_file_id", ref)
            .single()
            .execute()
        )
        doc_row = existing.data
    document_id = doc_row["id"]

    # Wipe any previous chunks for this doc — full reindex.
    supabase.table("document_chunks").delete().eq(
        "document_id", document_id
    ).execute()

    # Keep the original bytes so this file can later be attached to an email
    # (see send_followup helpers in agent_tools.py) — Telegram file URLs
    # expire and web uploads have no other persistent copy, unlike Drive/
    # OneDrive files which get re-downloaded live via drive_file_id instead.
    storage_path = _store_original_bytes(supabase, user_id=user_id, document_id=document_id, data=data)
    if storage_path:
        supabase.table("drive_documents").update({"storage_path": storage_path}).eq(
            "id", document_id
        ).execute()

    try:
        text = await _extract_text_from_blob(
            data, mime_type, name, genai_client=genai_client, user_id=user_id
        )
        if not text.strip():
            m_lower = (mime_type or "").lower()
            name_lower = name.lower()
            is_pdf = m_lower == "application/pdf" or name_lower.endswith(".pdf")
            is_image = m_lower.startswith("image/") or name_lower.endswith(
                (".jpg", ".jpeg", ".png", ".webp", ".heic", ".heif")
            )
            if is_pdf:
                err = (
                    "OCR ran on this scanned PDF but recovered no readable text. "
                    "Pages may be blank or in an unreadable script. Try a "
                    "different copy of the file."
                )
            elif is_image:
                err = (
                    "OCR ran on this image but found no readable text — it may "
                    "be a photo with no text in it (a face, a scene, etc.), "
                    "rather than a document/screenshot."
                )
            else:
                err = "no extractable text from this file type."
            supabase.table("drive_documents").update(
                {
                    "status": "failed",
                    "error": err,
                    "chunk_count": 0,
                }
            ).eq("id", document_id).execute()
            return {**doc_row, "status": "failed", "error": err}

        chunks = chunk_text(text)
        rows: list[dict[str, Any]] = []
        if chunks:
            titles = [name] * len(chunks)
            try:
                vectors = mem.embed_documents_batch(
                    genai_client, chunks, titles=titles, batch_size=50
                )
            except Exception as exc:  # noqa: BLE001
                logger.warning("blob batch embed failed for %s: %s", name, exc)
                vectors = []
                for c in chunks:
                    try:
                        vectors.append(mem.embed_document(genai_client, c, title=name))
                    except Exception:  # noqa: BLE001
                        vectors.append(None)
            for idx, (c, vec) in enumerate(zip(chunks, vectors)):
                if vec is None:
                    continue
                rows.append({
                    "user_id": user_id,
                    "document_id": document_id,
                    "drive_file_id": ref,
                    "file_name": name,
                    "chunk_index": idx,
                    "content": c,
                    "embedding": vec,
                    "metadata": {"mime_type": mime_type, "source": source},
                })
        if rows:
            for batch_start in range(0, len(rows), 50):
                supabase.table("document_chunks").insert(
                    rows[batch_start : batch_start + 50]
                ).execute()
            final_status = "indexed"
            err_msg = None
        else:
            final_status = "failed"
            err_msg = "embeddings failed for all chunks"

        supabase.table("drive_documents").update(
            {
                "status": final_status,
                "error": err_msg,
                "chunk_count": len(rows),
                "indexed_at": "now()",
            }
        ).eq("id", document_id).execute()
        return {**doc_row, "status": final_status, "chunk_count": len(rows)}

    except Exception as exc:  # noqa: BLE001
        logger.exception("index_blob failed for %s", name)
        supabase.table("drive_documents").update(
            {"status": "failed", "error": str(exc)[:500]}
        ).eq("id", document_id).execute()
        return {**doc_row, "status": "failed", "error": str(exc)}


async def index_folder(
    supabase,
    genai_client: genai.Client,
    *,
    user: dict[str, Any],
    folder_id: str,
    max_files: int = 50,
    source: str = "gdrive",
) -> list[dict[str, Any]]:
    """Walk a Drive/OneDrive folder (NOT recursive) and index every supported file."""
    if source == "onedrive":
        listing = await ms.list_onedrive_files(
            supabase, user, folder_id=folder_id, page_size=max_files
        )
        files = [
            f
            for f in listing.get("files", [])
            if not f.get("is_folder")
            and f.get("mime_type") in ms.ONEDRIVE_SUPPORTED_MIMES
        ]
    else:
        indexable_mimes = list(g.NATIVE_EXPORT.keys()) + list(g.SUPPORTED_BINARY_MIMES)
        listing = await g.list_drive_files(
            supabase,
            user,
            folder_id=folder_id,
            page_size=max_files,
            mime_filter=indexable_mimes,
        )
        files = listing.get("files", [])

    out: list[dict[str, Any]] = []
    for f in files:
        try:
            row = await index_file(
                supabase,
                genai_client,
                user=user,
                file_id=f["id"],
                file_meta=f,
                source=source,
            )
            out.append(row)
        except Exception:  # noqa: BLE001
            logger.exception("index_file crashed for %s", f.get("id"))
    return out


# ---------------------------------------------------------------------------
# Search
# ---------------------------------------------------------------------------


def search(
    supabase,
    genai_client: genai.Client,
    *,
    user_id: str,
    query: str,
    count: int = 8,
    threshold: float = 0.40,
) -> list[dict[str, Any]]:
    try:
        vec = mem.embed_query(genai_client, query)
    except Exception:  # noqa: BLE001
        logger.exception("embed_query failed")
        return []
    try:
        res = supabase.rpc(
            "match_document_chunks",
            {
                "query_embedding": vec,
                "match_user_id": user_id,
                "match_threshold": threshold,
                "match_count": count,
            },
        ).execute()
        return res.data or []
    except Exception:  # noqa: BLE001
        logger.exception("match_document_chunks rpc failed")
        return []


def format_for_prompt(chunks: list[dict[str, Any]]) -> str:
    if not chunks:
        return ""
    lines = ["Document excerpts (use these to ground your answer; cite filenames):"]
    for c in chunks:
        lines.append(
            f"\n--- {c.get('file_name')} (chunk {c.get('chunk_index')}, "
            f"similarity {c.get('similarity', 0):.2f}) ---\n{c.get('content', '')}"
        )
    return "\n".join(lines)
